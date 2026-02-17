"""
Excelレポート生成クラス
プロジェクトの比較結果を詳細なExcelレポートとして出力
"""
from typing import List, Dict, Optional
from pathlib import Path
import difflib
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from PIL import Image
import io
import os


class ReportGenerator:
    """Excelレポート生成クラス"""
    
    def __init__(self):
        """初期化"""
        self.workbook = None
        self.worksheet = None
    
    def generate_excel_report(
        self,
        output_path: str,
        web_pages: List,
        pdf_pages: List,
        pairs: List,
        project_name: str = "比較プロジェクト"
    ) -> bool:
        """
        Excelレポートを生成
        
        Args:
            output_path: 出力ファイルパス
            web_pages: Webページリスト
            pdf_pages: PDFページリスト
            pairs: マッチングペアリスト
            project_name: プロジェクト名
        
        Returns:
            bool: 成功時True
        """
        try:
            # ワークブック作成
            self.workbook = Workbook()
            self.worksheet = self.workbook.active
            self.worksheet.title = "比較結果"
            
            # ヘッダー行を設定
            self._setup_header(project_name)
            
            # データ行を追加
            current_row = 3
            for pair in pairs:
                web_page = self._find_page_by_id(web_pages, pair.web_id)
                pdf_page = self._find_page_by_id(pdf_pages, pair.pdf_id)
                
                if web_page and pdf_page:
                    current_row = self._add_comparison_row(
                        current_row,
                        web_page,
                        pdf_page,
                        pair.score
                    )
            
            # 列幅を自動調整
            self._adjust_column_widths()
            
            # ファイル保存
            self.workbook.save(output_path)
            print(f"✅ Excelレポートを出力しました: {output_path}")
            return True
            
        except Exception as e:
            print(f"⚠️ Excelレポート生成エラー: {e}")
            import traceback
            traceback.print_exc()
            return False
    
    def _setup_header(self, project_name: str):
        """ヘッダー行を設定"""
        # タイトル行
        self.worksheet.merge_cells('A1:G1')
        title_cell = self.worksheet['A1']
        title_cell.value = f"📊 {project_name} - 比較レポート"
        title_cell.font = Font(size=16, bold=True, color="FFFFFF")
        title_cell.fill = PatternFill(start_color="2196F3", end_color="2196F3", fill_type="solid")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        self.worksheet.row_dimensions[1].height = 30
        
        # カラムヘッダー
        headers = [
            "No.",
            "シンクロ率",
            "Web画像",
            "Webテキスト",
            "PDF画像",
            "PDFテキスト",
            "備考"
        ]
        
        for col_num, header in enumerate(headers, start=1):
            cell = self.worksheet.cell(row=2, column=col_num)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = self._get_border()
        
        self.worksheet.row_dimensions[2].height = 25
    
    def _add_comparison_row(
        self,
        row_num: int,
        web_page,
        pdf_page,
        score: float
    ) -> int:
        """
        比較データ行を追加
        
        Args:
            row_num: 行番号
            web_page: Webページオブジェクト
            pdf_page: PDFページオブジェクト
            score: 類似度スコア
        
        Returns:
            int: 次の行番号
        """
        # 行の高さを設定（画像用）
        self.worksheet.row_dimensions[row_num].height = 200
        
        # A列: No.
        no_cell = self.worksheet.cell(row=row_num, column=1)
        no_cell.value = row_num - 2
        no_cell.alignment = Alignment(horizontal="center", vertical="center")
        no_cell.border = self._get_border()
        
        # B列: シンクロ率
        score_cell = self.worksheet.cell(row=row_num, column=2)
        score_cell.value = f"{int(score * 100)}%"
        score_cell.alignment = Alignment(horizontal="center", vertical="center")
        score_cell.border = self._get_border()
        
        # スコアに応じて色分け
        if score >= 0.7:
            score_cell.fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
            score_cell.font = Font(bold=True, color="2E7D32")
        elif score >= 0.4:
            score_cell.fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
            score_cell.font = Font(bold=True, color="F57F17")
        else:
            score_cell.fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
            score_cell.font = Font(bold=True, color="C62828")
        
        # C列: Web画像
        if web_page.screenshot_image:
            self._add_image_to_cell(
                web_page.screenshot_image,
                row_num,
                3,
                max_width=250,
                max_height=180
            )
        
        # D列: Webテキスト
        web_text_cell = self.worksheet.cell(row=row_num, column=4)
        web_text_cell.value = web_page.text if web_page.text else "(テキストなし)"
        web_text_cell.alignment = Alignment(
            horizontal="left",
            vertical="top",
            wrap_text=True
        )
        web_text_cell.border = self._get_border()
        
        # E列: PDF画像
        if pdf_page.page_image:
            self._add_image_to_cell(
                pdf_page.page_image,
                row_num,
                5,
                max_width=250,
                max_height=180
            )
        
        # F列: PDFテキスト（差分を強調）
        pdf_text_cell = self.worksheet.cell(row=row_num, column=6)
        pdf_text = pdf_page.text if pdf_page.text else "(テキストなし)"
        
        # テキスト差分を検出して強調（簡易版：セル全体の色で判定）
        if web_page.text and pdf_page.text:
            if web_page.text != pdf_page.text:
                # 差分がある場合、セルの背景を薄いピンクに
                pdf_text_cell.fill = PatternFill(start_color="FFE0E0", end_color="FFE0E0", fill_type="solid")
        
        pdf_text_cell.value = pdf_text
        pdf_text_cell.alignment = Alignment(
            horizontal="left",
            vertical="top",
            wrap_text=True
        )
        pdf_text_cell.border = self._get_border()
        
        # G列: 備考（URL、ファイル名）
        remarks = f"Web: {web_page.url}\n\nPDF: {Path(pdf_page.filename).name}\nページ: {pdf_page.page_num}"
        remarks_cell = self.worksheet.cell(row=row_num, column=7)
        remarks_cell.value = remarks
        remarks_cell.alignment = Alignment(
            horizontal="left",
            vertical="top",
            wrap_text=True
        )
        remarks_cell.border = self._get_border()
        
        return row_num + 1
    
    def _add_image_to_cell(
        self,
        pil_image: Image.Image,
        row: int,
        col: int,
        max_width: int = 250,
        max_height: int = 180
    ):
        """
        PIL ImageをExcelセルに貼り付け
        
        Args:
            pil_image: PIL Imageオブジェクト
            row: 行番号
            col: 列番号
            max_width: 最大幅
            max_height: 最大高さ
        """
        try:
            # 画像をリサイズ（アスペクト比を保持）
            img_copy = pil_image.copy()
            img_copy.thumbnail((max_width, max_height), Image.Resampling.LANCZOS)
            
            # PIL ImageをバイトストリームでExcel画像に変換
            img_buffer = io.BytesIO()
            img_copy.save(img_buffer, format='PNG')
            img_buffer.seek(0)
            
            # Excel画像オブジェクトを作成
            xl_image = XLImage(img_buffer)
            
            # セルの位置を計算
            col_letter = get_column_letter(col)
            anchor_cell = f"{col_letter}{row}"
            
            # 画像を配置
            xl_image.anchor = anchor_cell
            self.worksheet.add_image(xl_image)
            
        except Exception as e:
            print(f"⚠️ 画像貼り付けエラー: {e}")
    
    def _adjust_column_widths(self):
        """列幅を調整"""
        column_widths = {
            1: 8,   # No.
            2: 12,  # シンクロ率
            3: 35,  # Web画像
            4: 50,  # Webテキスト
            5: 35,  # PDF画像
            6: 50,  # PDFテキスト
            7: 30   # 備考
        }
        
        for col, width in column_widths.items():
            col_letter = get_column_letter(col)
            self.worksheet.column_dimensions[col_letter].width = width
    
    def _get_border(self) -> Border:
        """セルの枠線を取得"""
        thin_border = Side(border_style="thin", color="CCCCCC")
        return Border(
            left=thin_border,
            right=thin_border,
            top=thin_border,
            bottom=thin_border
        )
    
    def _find_page_by_id(self, pages: List, page_id: int):
        """IDからページを検索"""
        for page in pages:
            if page.page_id == page_id:
                return page
        return None
    
    def generate_detailed_diff_report(
        self,
        output_path: str,
        web_pages: List,
        pdf_pages: List,
        pairs: List,
        project_name: str = "比較プロジェクト"
    ) -> bool:
        """
        詳細な差分レポートを生成（テキスト差分を行単位で強調）
        
        Args:
            output_path: 出力ファイルパス
            web_pages: Webページリスト
            pdf_pages: PDFページリスト
            pairs: マッチングペアリスト
            project_name: プロジェクト名
        
        Returns:
            bool: 成功時True
        """
        try:
            # 基本レポートを生成
            success = self.generate_excel_report(
                output_path,
                web_pages,
                pdf_pages,
                pairs,
                project_name
            )
            
            if success:
                # 差分シートを追加
                diff_sheet = self.workbook.create_sheet("詳細差分")
                self._add_diff_analysis(diff_sheet, web_pages, pdf_pages, pairs)
                self.workbook.save(output_path)
            
            return success
            
        except Exception as e:
            print(f"⚠️ 詳細レポート生成エラー: {e}")
            return False
    
    def _add_diff_analysis(self, sheet, web_pages, pdf_pages, pairs):
        """差分分析シートを追加"""
        # ヘッダー
        sheet['A1'] = "No."
        sheet['B1'] = "Web/PDF"
        sheet['C1'] = "行番号"
        sheet['D1'] = "テキスト"
        sheet['E1'] = "状態"
        
        for col in range(1, 6):
            cell = sheet.cell(row=1, column=col)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        row_num = 2
        pair_num = 1
        
        for pair in pairs:
            web_page = self._find_page_by_id(web_pages, pair.web_id)
            pdf_page = self._find_page_by_id(pdf_pages, pair.pdf_id)
            
            if web_page and pdf_page:
                # テキストを行単位で比較
                web_lines = web_page.text.split('\n') if web_page.text else []
                pdf_lines = pdf_page.text.split('\n') if pdf_page.text else []
                
                # difflibで差分を取得
                diff = difflib.ndiff(web_lines, pdf_lines)
                
                for line in diff:
                    status = line[0]
                    text = line[2:]
                    
                    sheet.cell(row=row_num, column=1).value = pair_num
                    
                    if status == ' ':
                        # 一致
                        sheet.cell(row=row_num, column=2).value = "共通"
                        sheet.cell(row=row_num, column=4).value = text
                        sheet.cell(row=row_num, column=5).value = "一致"
                        sheet.cell(row=row_num, column=5).fill = PatternFill(
                            start_color="C8E6C9", end_color="C8E6C9", fill_type="solid"
                        )
                    elif status == '-':
                        # Webのみ
                        sheet.cell(row=row_num, column=2).value = "Web"
                        sheet.cell(row=row_num, column=4).value = text
                        sheet.cell(row=row_num, column=5).value = "削除"
                        sheet.cell(row=row_num, column=5).fill = PatternFill(
                            start_color="FFCDD2", end_color="FFCDD2", fill_type="solid"
                        )
                        sheet.cell(row=row_num, column=4).font = Font(color="C62828")
                    elif status == '+':
                        # PDFのみ
                        sheet.cell(row=row_num, column=2).value = "PDF"
                        sheet.cell(row=row_num, column=4).value = text
                        sheet.cell(row=row_num, column=5).value = "追加"
                        sheet.cell(row=row_num, column=5).fill = PatternFill(
                            start_color="BBDEFB", end_color="BBDEFB", fill_type="solid"
                        )
                        sheet.cell(row=row_num, column=4).font = Font(color="1976D2")
                    
                    row_num += 1
                
                pair_num += 1
        
        # 列幅調整
        sheet.column_dimensions['A'].width = 8
        sheet.column_dimensions['B'].width = 12
        sheet.column_dimensions['C'].width = 10
        sheet.column_dimensions['D'].width = 80
        sheet.column_dimensions['E'].width = 12

