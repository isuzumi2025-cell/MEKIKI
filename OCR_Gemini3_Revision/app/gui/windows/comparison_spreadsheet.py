"""
æ¯”è¼ƒã‚¹ãƒ—ãƒ¬ãƒƒãƒ‰ã‚·ãƒ¼ãƒˆã‚¦ã‚£ãƒ³ãƒ‰ã‚¦ (ç”»é¢2) - 2åˆ—æ§‹æˆç‰ˆ
Web/PDFæ¨ªä¸¦ã³ã€ã‚·ãƒ³ã‚¯ãƒ­ç®‡æ‰€ã¯åŒåˆ—ã«é…ç½®
"""

import customtkinter as ctk
import tkinter as tk
from tkinter import ttk
from typing import List, Dict, Optional, Callable, Tuple
from PIL import Image, ImageTk
from dataclasses import dataclass
import io


@dataclass  
class ComparisonRow:
    """æ¯”è¼ƒè¡Œãƒ‡ãƒ¼ã‚¿ï¼ˆ2åˆ—æ§‹æˆï¼‰"""
    row_no: int
    # Webå´
    web_id: str = ""
    web_text: str = ""
    web_rect: List[int] = None
    # PDFå´
    pdf_id: str = ""
    pdf_text: str = ""
    pdf_rect: List[int] = None
    # ãƒãƒƒãƒæƒ…å ±
    similarity: float = 0.0
    sync_area: str = ""  # "P1-3 â†” PDF-5"
    
    @property
    def status_icon(self) -> str:
        if self.similarity >= 0.5:
            return "ğŸŸ¢"
        elif self.similarity >= 0.3:
            return "ğŸŸ¡"
        elif self.sync_area:
            return "ğŸ”´"
        else:
            return "âšª"
    
    @property
    def web_preview(self) -> str:
        text = self.web_text.replace('\n', ' ')[:60]
        return text + "..." if len(self.web_text) > 60 else text
    
    @property
    def pdf_preview(self) -> str:
        text = self.pdf_text.replace('\n', ' ')[:60]
        return text + "..." if len(self.pdf_text) > 60 else text


class ComparisonSpreadsheetWindow(ctk.CTkToplevel):
    """
    2åˆ—æ§‹æˆã®æ¯”è¼ƒã‚¹ãƒ—ãƒ¬ãƒƒãƒ‰ã‚·ãƒ¼ãƒˆ
    â”Œâ”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”¬â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”¬â”€â”€â”€â”€â”€â”€â”
    â”‚  Web (No/ç”»åƒ/ãƒ†ã‚­ã‚¹ãƒˆ) â”‚ PDF (No/ç”»åƒ/ãƒ†ã‚­ã‚¹ãƒˆ)  â”‚Sync  â”‚
    â””â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”´â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”´â”€â”€â”€â”€â”€â”€â”˜
    """
    
    def __init__(self, parent, on_row_select: Callable = None, **kwargs):
        super().__init__(parent)
        
        self.title("ğŸ“Š Web/PDF æ¯”è¼ƒæ ¡æ­£ã‚·ãƒ¼ãƒˆ (2åˆ—æ§‹æˆ)")
        self.geometry("1400x750")
        self.configure(fg_color="#1A1A1A")
        
        self.on_row_select = on_row_select
        self.rows: List[ComparisonRow] = []
        self.web_image: Optional[Image.Image] = None
        self.pdf_image: Optional[Image.Image] = None
        self.thumbnails: Dict[int, Tuple[ImageTk.PhotoImage, ImageTk.PhotoImage]] = {}
        
        self._build_ui()
    
    def _build_ui(self):
        """UIæ§‹ç¯‰"""
        # ãƒ˜ãƒƒãƒ€ãƒ¼
        header = ctk.CTkFrame(self, fg_color="#2D2D2D", height=50)
        header.pack(fill="x")
        header.pack_propagate(False)
        
        ctk.CTkLabel(
            header,
            text="ğŸ“Š Web/PDF æ¯”è¼ƒæ ¡æ­£ã‚·ãƒ¼ãƒˆ",
            font=("Meiryo", 14, "bold")
        ).pack(side="left", padx=15, pady=10)
        
        self.count_label = ctk.CTkLabel(
            header, text="0ä»¶", font=("Meiryo", 11), text_color="gray"
        )
        self.count_label.pack(side="left", padx=10)
        
        # ãƒ„ãƒ¼ãƒ«ãƒãƒ¼
        toolbar = ctk.CTkFrame(header, fg_color="transparent")
        toolbar.pack(side="right", padx=10)
        
        ctk.CTkButton(
            toolbar, text="ğŸ“¥ Excelå‡ºåŠ›", width=100, fg_color="#4CAF50",
            command=self._export_excel
        ).pack(side="left", padx=5)
        
        ctk.CTkButton(
            toolbar, text="ğŸ”„ æ›´æ–°", width=80, fg_color="#2196F3",
            command=self._refresh
        ).pack(side="left", padx=5)
        
        # ã‚«ãƒ©ãƒ ãƒ˜ãƒƒãƒ€ãƒ¼ (2åˆ—æ§‹æˆ)
        col_header = ctk.CTkFrame(self, fg_color="#383838", height=40)
        col_header.pack(fill="x", padx=5, pady=2)
        col_header.pack_propagate(False)
        
        # Webåˆ—ãƒ˜ãƒƒãƒ€ãƒ¼
        web_header = ctk.CTkFrame(col_header, fg_color="#2E7D32", width=600)
        web_header.pack(side="left", fill="y", padx=1)
        web_header.pack_propagate(False)
        ctk.CTkLabel(web_header, text="ğŸŒ Webå´", font=("Meiryo", 12, "bold")).pack(pady=8)
        
        # PDFåˆ—ãƒ˜ãƒƒãƒ€ãƒ¼
        pdf_header = ctk.CTkFrame(col_header, fg_color="#1565C0", width=600)
        pdf_header.pack(side="left", fill="y", padx=1)
        pdf_header.pack_propagate(False)
        ctk.CTkLabel(pdf_header, text="ğŸ“„ PDFå´", font=("Meiryo", 12, "bold")).pack(pady=8)
        
        # Syncåˆ—ãƒ˜ãƒƒãƒ€ãƒ¼
        sync_header = ctk.CTkFrame(col_header, fg_color="#FF6F00", width=180)
        sync_header.pack(side="left", fill="y", padx=1)
        sync_header.pack_propagate(False)
        ctk.CTkLabel(sync_header, text="ğŸ”— Sync", font=("Meiryo", 12, "bold")).pack(pady=8)
        
        # ã‚µãƒ–ãƒ˜ãƒƒãƒ€ãƒ¼
        sub_header = ctk.CTkFrame(self, fg_color="#2D2D2D", height=25)
        sub_header.pack(fill="x", padx=5)
        sub_header.pack_propagate(False)
        
        # Webå´ã‚µãƒ–ãƒ˜ãƒƒãƒ€ãƒ¼
        for text, w in [("No", 50), ("ç”»åƒ", 100), ("ãƒ†ã‚­ã‚¹ãƒˆ", 430)]:
            ctk.CTkLabel(sub_header, text=text, width=w, font=("Meiryo", 9)).pack(side="left", padx=1)
        
        # ã‚¹ãƒšãƒ¼ã‚µãƒ¼
        ctk.CTkLabel(sub_header, text="|", width=10, text_color="gray").pack(side="left")
        
        # PDFå´ã‚µãƒ–ãƒ˜ãƒƒãƒ€ãƒ¼
        for text, w in [("No", 50), ("ç”»åƒ", 100), ("ãƒ†ã‚­ã‚¹ãƒˆ", 430)]:
            ctk.CTkLabel(sub_header, text=text, width=w, font=("Meiryo", 9)).pack(side="left", padx=1)
        
        # ã‚¹ãƒšãƒ¼ã‚µãƒ¼
        ctk.CTkLabel(sub_header, text="|", width=10, text_color="gray").pack(side="left")
        
        # Syncå´ã‚µãƒ–ãƒ˜ãƒƒãƒ€ãƒ¼
        for text, w in [("ç‡", 50), ("ã‚¨ãƒªã‚¢", 120)]:
            ctk.CTkLabel(sub_header, text=text, width=w, font=("Meiryo", 9)).pack(side="left", padx=1)
        
        # ãƒ¡ã‚¤ãƒ³ãƒ•ãƒ¬ãƒ¼ãƒ 
        main_frame = ctk.CTkFrame(self, fg_color="#1E1E1E")
        main_frame.pack(fill="both", expand=True, padx=5, pady=5)
        
        self.canvas = tk.Canvas(main_frame, bg="#1E1E1E", highlightthickness=0)
        scrollbar_y = ttk.Scrollbar(main_frame, orient="vertical", command=self.canvas.yview)
        
        self.scrollable_frame = ctk.CTkFrame(self.canvas, fg_color="#1E1E1E")
        self.scrollable_frame.bind("<Configure>", lambda e: self.canvas.configure(scrollregion=self.canvas.bbox("all")))
        
        self.canvas.create_window((0, 0), window=self.scrollable_frame, anchor="nw")
        self.canvas.configure(yscrollcommand=scrollbar_y.set)
        
        scrollbar_y.pack(side="right", fill="y")
        self.canvas.pack(side="left", fill="both", expand=True)
        
        self.canvas.bind_all("<MouseWheel>", self._on_mousewheel)
    
    def load_data(self, web_regions: List, pdf_regions: List, 
                  web_image: Image.Image = None, pdf_image: Image.Image = None,
                  sync_pairs: List = None):
        """ãƒ‡ãƒ¼ã‚¿ã‚’ãƒ­ãƒ¼ãƒ‰ï¼ˆ2åˆ—æ§‹æˆç”¨ï¼‰"""
        self.web_image = web_image
        self.pdf_image = pdf_image
        self.rows = []
        
        web_map = {r.area_code: r for r in web_regions}
        pdf_map = {r.area_code: r for r in pdf_regions}
        
        used_web = set()
        used_pdf = set()
        row_no = 1
        
        # 1. ãƒãƒƒãƒãƒšã‚¢ã‚’è¡Œã«ï¼ˆæ¨ªä¸¦ã³ï¼‰
        if sync_pairs:
            for sp in sync_pairs:
                web_r = web_map.get(sp.web_id)
                pdf_r = pdf_map.get(sp.pdf_id)
                
                row = ComparisonRow(
                    row_no=row_no,
                    web_id=sp.web_id,
                    web_text=web_r.text if web_r else "",
                    web_rect=list(web_r.rect) if web_r else None,
                    pdf_id=sp.pdf_id,
                    pdf_text=pdf_r.text if pdf_r else "",
                    pdf_rect=list(pdf_r.rect) if pdf_r else None,
                    similarity=sp.similarity,
                    sync_area=f"{sp.web_id} â†” {sp.pdf_id}"
                )
                self.rows.append(row)
                used_web.add(sp.web_id)
                used_pdf.add(sp.pdf_id)
                row_no += 1
        
        # 2. ãƒãƒƒãƒã—ãªã‹ã£ãŸWeb (PDFå´ç©ºæ¬„)
        for r in web_regions:
            if r.area_code not in used_web:
                row = ComparisonRow(
                    row_no=row_no,
                    web_id=r.area_code,
                    web_text=r.text,
                    web_rect=list(r.rect),
                    similarity=0.0
                )
                self.rows.append(row)
                row_no += 1
        
        # 3. ãƒãƒƒãƒã—ãªã‹ã£ãŸPDF (Webå´ç©ºæ¬„)
        for r in pdf_regions:
            if r.area_code not in used_pdf:
                row = ComparisonRow(
                    row_no=row_no,
                    pdf_id=r.area_code,
                    pdf_text=r.text,
                    pdf_rect=list(r.rect),
                    similarity=0.0
                )
                self.rows.append(row)
                row_no += 1
        
        self._generate_thumbnails()
        self._refresh_rows()
    
    def _generate_thumbnails(self):
        """ã‚µãƒ ãƒã‚¤ãƒ«ç”Ÿæˆ"""
        self.thumbnails = {}
        thumb_h = 50
        
        for row in self.rows:
            web_thumb = None
            pdf_thumb = None
            
            if self.web_image and row.web_rect:
                try:
                    x1, y1, x2, y2 = row.web_rect
                    cropped = self.web_image.crop((max(0,x1), max(0,y1), min(self.web_image.width,x2), min(self.web_image.height,y2)))
                    if cropped.height > 0:
                        ratio = thumb_h / cropped.height
                        resized = cropped.resize((min(int(cropped.width * ratio), 90), thumb_h), Image.Resampling.LANCZOS)
                        web_thumb = ImageTk.PhotoImage(resized)
                except: pass
            
            if self.pdf_image and row.pdf_rect:
                try:
                    x1, y1, x2, y2 = row.pdf_rect
                    cropped = self.pdf_image.crop((max(0,x1), max(0,y1), min(self.pdf_image.width,x2), min(self.pdf_image.height,y2)))
                    if cropped.height > 0:
                        ratio = thumb_h / cropped.height
                        resized = cropped.resize((min(int(cropped.width * ratio), 90), thumb_h), Image.Resampling.LANCZOS)
                        pdf_thumb = ImageTk.PhotoImage(resized)
                except: pass
            
            self.thumbnails[row.row_no] = (web_thumb, pdf_thumb)
    
    def _refresh_rows(self):
        """è¡Œã‚’å†æç”»"""
        for widget in self.scrollable_frame.winfo_children():
            widget.destroy()
        
        for i, row in enumerate(self.rows):
            self._create_row(row, i)
        
        self.count_label.configure(text=f"{len(self.rows)}ä»¶")
    
    def _create_row(self, row: ComparisonRow, index: int):
        """1è¡Œä½œæˆï¼ˆ2åˆ—æ§‹æˆï¼‰"""
        bg = "#2B2B2B" if index % 2 == 0 else "#333333"
        row_frame = ctk.CTkFrame(self.scrollable_frame, fg_color=bg, height=60)
        row_frame.pack(fill="x", pady=1)
        row_frame.pack_propagate(False)
        
        row_frame.bind("<Button-1>", lambda e, r=row: self._on_row_click(r))
        row_frame.bind("<Double-1>", lambda e, r=row: self._on_row_double_click(r))
        
        # === Webåˆ— ===
        # No
        ctk.CTkLabel(row_frame, text=row.web_id or "-", width=50, font=("Meiryo", 9), text_color="#4CAF50").pack(side="left", padx=1)
        
        # ç”»åƒ
        web_thumb, pdf_thumb = self.thumbnails.get(row.row_no, (None, None))
        web_img_frame = ctk.CTkFrame(row_frame, fg_color="#1E1E1E", width=100, height=55)
        web_img_frame.pack(side="left", padx=1)
        web_img_frame.pack_propagate(False)
        if web_thumb:
            lbl = tk.Label(web_img_frame, image=web_thumb, bg="#1E1E1E")
            lbl.image = web_thumb
            lbl.pack(expand=True)
        
        # ãƒ†ã‚­ã‚¹ãƒˆ
        ctk.CTkLabel(row_frame, text=row.web_preview, width=430, font=("Meiryo", 9), anchor="w").pack(side="left", padx=1)
        
        # åŒºåˆ‡ã‚Š
        ctk.CTkLabel(row_frame, text="â”‚", width=10, text_color="#555").pack(side="left")
        
        # === PDFåˆ— ===
        # No
        ctk.CTkLabel(row_frame, text=row.pdf_id or "-", width=50, font=("Meiryo", 9), text_color="#2196F3").pack(side="left", padx=1)
        
        # ç”»åƒ
        pdf_img_frame = ctk.CTkFrame(row_frame, fg_color="#1E1E1E", width=100, height=55)
        pdf_img_frame.pack(side="left", padx=1)
        pdf_img_frame.pack_propagate(False)
        if pdf_thumb:
            lbl = tk.Label(pdf_img_frame, image=pdf_thumb, bg="#1E1E1E")
            lbl.image = pdf_thumb
            lbl.pack(expand=True)
        
        # ãƒ†ã‚­ã‚¹ãƒˆ
        ctk.CTkLabel(row_frame, text=row.pdf_preview, width=430, font=("Meiryo", 9), anchor="w").pack(side="left", padx=1)
        
        # åŒºåˆ‡ã‚Š
        ctk.CTkLabel(row_frame, text="â”‚", width=10, text_color="#555").pack(side="left")
        
        # === Syncåˆ— ===
        sim_color = "#4CAF50" if row.similarity >= 0.5 else "#FF9800" if row.similarity >= 0.3 else "#888"
        ctk.CTkLabel(row_frame, text=f"{row.status_icon} {row.similarity*100:.0f}%", width=50, font=("Meiryo", 9, "bold"), text_color=sim_color).pack(side="left", padx=1)
        ctk.CTkLabel(row_frame, text=row.sync_area, width=120, font=("Meiryo", 8), text_color="gray").pack(side="left", padx=1)
    
    def _on_row_click(self, row):
        if self.on_row_select:
            self.on_row_select(row, "click")
    
    def _on_row_double_click(self, row):
        if self.on_row_select:
            self.on_row_select(row, "double_click")
    
    def _on_mousewheel(self, event):
        self.canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
    
    def _refresh(self):
        self._generate_thumbnails()
        self._refresh_rows()
    
    def _export_excel(self):
        """Excelå‡ºåŠ›ï¼ˆ2åˆ—æ§‹æˆï¼‰"""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
            from pathlib import Path
            from datetime import datetime
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "æ¯”è¼ƒã‚·ãƒ¼ãƒˆ"
            
            # ãƒ˜ãƒƒãƒ€ãƒ¼
            headers = ["No", "Web ID", "Webãƒ†ã‚­ã‚¹ãƒˆ", "PDF ID", "PDFãƒ†ã‚­ã‚¹ãƒˆ", "Syncç‡", "Syncã‚¨ãƒªã‚¢"]
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="4CAF50", fill_type="solid")
            
            # ãƒ‡ãƒ¼ã‚¿
            for i, row in enumerate(self.rows, 2):
                ws.cell(row=i, column=1, value=row.row_no)
                ws.cell(row=i, column=2, value=row.web_id)
                ws.cell(row=i, column=3, value=row.web_text[:200])
                ws.cell(row=i, column=4, value=row.pdf_id)
                ws.cell(row=i, column=5, value=row.pdf_text[:200])
                ws.cell(row=i, column=6, value=f"{row.similarity*100:.0f}%")
                ws.cell(row=i, column=7, value=row.sync_area)
            
            # åˆ—å¹…èª¿æ•´
            ws.column_dimensions['C'].width = 50
            ws.column_dimensions['E'].width = 50
            
            # ä¿å­˜
            Path("./exports").mkdir(exist_ok=True)
            filename = f"comparison_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            output_path = f"./exports/{filename}"
            wb.save(output_path)
            
            print(f"âœ… Excelå‡ºåŠ›å®Œäº†: {output_path}")
            import os
            os.startfile(output_path)
            
        except Exception as e:
            print(f"Excelå‡ºåŠ›ã‚¨ãƒ©ãƒ¼: {e}")
            import traceback
            traceback.print_exc()
