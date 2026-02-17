import customtkinter as ctk
import tkinter as tk
from typing import List, Any, Callable, Optional
from PIL import Image, ImageTk
import difflib

class SpreadsheetPanel(ctk.CTkFrame):
    """
    Live Comparison Sheet (Virtual Spreadsheet) - Phase 5 Enhanced
    - Thumbnails below ID (click to jump to Source)
    - Auto-expanding row heights
    - Diff coloring for text comparison
    """
    def __init__(self, parent, on_row_select: Optional[Callable] = None, **kwargs):
        super().__init__(parent, **kwargs)
        
        # Data reference
        self.sync_pairs = []
        self.web_map = {}
        self.pdf_map = {}
        self.web_image = None
        self.pdf_image = None
        self.selected_indices = None
        self.selected_widget = None
        self.on_row_select = on_row_select
        self._thumbnail_refs = []  # Keep references to prevent GC
        
        self._build_ui()
        
    def set_on_row_select(self, callback: Callable):
        """Set callback for row selection"""
        self.on_row_select = callback
        
    def set_images(self, web_image: Image.Image, pdf_image: Image.Image):
        """Set source images for thumbnail generation"""
        self.web_image = web_image
        self.pdf_image = pdf_image
        
    def _build_ui(self):
        # 1. Toolbar
        toolbar = ctk.CTkFrame(self, height=30, fg_color="#333333")
        toolbar.pack(fill="x", side="top")
        
        ctk.CTkLabel(toolbar, text="Live Comparison Sheet", font=("Meiryo", 12, "bold")).pack(side="left", padx=10)
        
        self.stats_label = ctk.CTkLabel(toolbar, text="Web: - | PDF: - | Match: -", font=("Meiryo", 11))
        self.stats_label.pack(side="left", padx=20)
        
        self.export_btn = ctk.CTkButton(
            toolbar, 
            text="Excel Export", 
            width=100, 
            height=24, 
            state="disabled",
            command=self._on_export
        )
        self.export_btn.pack(side="right", padx=5)
        
        # 2. Table Header (fixed)
        header_frame = ctk.CTkFrame(self, height=28, fg_color="#2B2B2B")
        header_frame.pack(fill="x", side="top", pady=(1, 0))
        
        # Header columns
        headers = [
            ("Score", 50),
            ("Web ID / Thumb", 100),
            ("Web Text", 0),
            ("", 30),
            ("PDF Text", 0),
            ("PDF ID / Thumb", 100),
        ]
        
        for text, width in headers:
            if width > 0:
                lbl = ctk.CTkLabel(header_frame, text=text, width=width, font=("Meiryo", 9, "bold"))
            else:
                lbl = ctk.CTkLabel(header_frame, text=text, font=("Meiryo", 9, "bold"))
            lbl.pack(side="left", padx=2, pady=4)
            if width == 0:
                lbl.pack_configure(expand=True, fill="x")
        
        # 3. Scrollable content area
        self.scroll_frame = ctk.CTkScrollableFrame(self, fg_color="#1E1E1E", corner_radius=0)
        self.scroll_frame.pack(fill="both", expand=True)
        
    def update_data(self, sync_pairs: List[Any], web_regions: List[Any], pdf_regions: List[Any], 
                    web_image=None, pdf_image=None):
        """Update data and refresh display"""
        self.sync_pairs = sync_pairs
        self.web_map = {r.area_code: r for r in web_regions if hasattr(r, 'area_code')}
        self.pdf_map = {r.area_code: r for r in pdf_regions if hasattr(r, 'area_code')}
        
        if web_image:
            self.web_image = web_image
        if pdf_image:
            self.pdf_image = pdf_image
        
        # Stats update (★ threshold_low=0.25)
        total_web = len(web_regions)
        total_pdf = len(pdf_regions)
        matched = sum(1 for p in sync_pairs if p.similarity >= 0.25)
        self.stats_label.configure(text=f"Web: {total_web} | PDF: {total_pdf} | Match: {matched}")
        
        # Refresh rows
        self._refresh_rows()
        
        # Enable export button
        if sync_pairs:
            self.export_btn.configure(state="normal")
            
    def _refresh_rows(self):
        # Clear existing rows
        for widget in self.scroll_frame.winfo_children():
            widget.destroy()
        self._thumbnail_refs = []
            
        # Add rows
        for i, pair in enumerate(self.sync_pairs):
            self._create_row(i, pair)
            
    def _create_row(self, index: int, pair):
        """Create a single row with thumbnails below ID"""
        row_bg = "#2D2D2D" if index % 2 == 0 else "#252525"
        
        row = ctk.CTkFrame(self.scroll_frame, fg_color=row_bg, corner_radius=0, height=120)
        row.pack(fill="x", pady=1)
        row.pack_propagate(False)
        
        # Get regions
        web_region = self.web_map.get(pair.web_id)
        pdf_region = self.pdf_map.get(pair.pdf_id)
        w_txt = web_region.text if web_region else ""
        p_txt = pdf_region.text if pdf_region else ""
        
        # === Score column ===
        sim_percent = int(pair.similarity * 100)
        if sim_percent >= 80:
            color, bg = "#4CAF50", "#1B3D1B"
        elif sim_percent >= 50:
            color, bg = "#FFC107", "#3D3D1B"
        elif sim_percent >= 30:
            color, bg = "#FF9800", "#3D2D1B"
        else:
            color, bg = "#F44336", "#3D1B1B"
            
        score_frame = ctk.CTkFrame(row, fg_color=bg, width=50)
        score_frame.pack(side="left", fill="y", padx=1)
        score_frame.pack_propagate(False)
        ctk.CTkLabel(score_frame, text=f"{sim_percent}%", text_color=color, font=("Arial", 10, "bold")).pack(expand=True)
        
        # === Web ID + Thumbnail column ===
        web_id_frame = ctk.CTkFrame(row, fg_color=row_bg, width=100)
        web_id_frame.pack(side="left", fill="y", padx=1)
        web_id_frame.pack_propagate(False)
        
        ctk.CTkLabel(web_id_frame, text=pair.web_id or "-", text_color="#4CAF50", font=("Meiryo", 8)).pack(pady=(2, 0))
        
        # Web thumbnail
        web_thumb = self._create_thumbnail(self.web_image, web_region, 45, 35)
        if web_thumb:
            self._thumbnail_refs.append(web_thumb)
            thumb_label = tk.Label(web_id_frame, image=web_thumb, bg=row_bg, cursor="hand2")
            thumb_label.pack(pady=2)
            thumb_label.bind("<Button-1>", lambda e, r=web_region: self._on_thumbnail_click(r, "web"))
        
        # === Web Text column ===
        web_text_frame = ctk.CTkFrame(row, fg_color=row_bg)
        web_text_frame.pack(side="left", fill="both", expand=True, padx=2)
        
        web_text = tk.Text(web_text_frame, bg=row_bg, fg="#E0E0E0", relief="flat",
                          font=("Meiryo", 18), wrap="word", height=5, width=25)
        web_text.pack(fill="both", expand=True, padx=2, pady=2)
        web_text.insert("1.0", w_txt[:200] + ("..." if len(w_txt) > 200 else ""))
        web_text.configure(state="disabled")
        
        # === Arrow column ===
        ctk.CTkLabel(row, text="<>", width=30, text_color="#666666").pack(side="left", padx=1)
        
        # === PDF Text column ===
        pdf_text_frame = ctk.CTkFrame(row, fg_color=row_bg)
        pdf_text_frame.pack(side="left", fill="both", expand=True, padx=2)
        
        pdf_text = tk.Text(pdf_text_frame, bg=row_bg, fg="#E0E0E0", relief="flat",
                          font=("Meiryo", 18), wrap="word", height=5, width=25)
        pdf_text.pack(fill="both", expand=True, padx=2, pady=2)
        pdf_text.insert("1.0", p_txt[:200] + ("..." if len(p_txt) > 200 else ""))
        pdf_text.configure(state="disabled")
        
        # === PDF ID + Thumbnail column ===
        pdf_id_frame = ctk.CTkFrame(row, fg_color=row_bg, width=100)
        pdf_id_frame.pack(side="left", fill="y", padx=1)
        pdf_id_frame.pack_propagate(False)
        
        ctk.CTkLabel(pdf_id_frame, text=pair.pdf_id or "-", text_color="#2196F3", font=("Meiryo", 8)).pack(pady=(2, 0))
        
        # PDF thumbnail
        pdf_thumb = self._create_thumbnail(self.pdf_image, pdf_region, 45, 35)
        if pdf_thumb:
            self._thumbnail_refs.append(pdf_thumb)
            thumb_label = tk.Label(pdf_id_frame, image=pdf_thumb, bg=row_bg, cursor="hand2")
            thumb_label.pack(pady=2)
            thumb_label.bind("<Button-1>", lambda e, r=pdf_region: self._on_thumbnail_click(r, "pdf"))
        
        # Row click binding
        def on_click(event, p=pair, w=row):
            self._on_row_click(w, p)
            
        row.bind("<Button-1>", on_click)
        for widget in [score_frame, web_id_frame, pdf_id_frame]:
            widget.bind("<Button-1>", on_click)
            
    def _create_thumbnail(self, source_image, region, width: int, height: int):
        """Create a thumbnail from the source image and region rect"""
        if not source_image or not region or not hasattr(region, 'rect'):
            return None
            
        try:
            x1, y1, x2, y2 = region.rect
            # Ensure valid coordinates
            x1 = max(0, int(x1))
            y1 = max(0, int(y1))
            x2 = min(source_image.width, int(x2))
            y2 = min(source_image.height, int(y2))
            
            if x2 <= x1 or y2 <= y1:
                return None
                
            cropped = source_image.crop((x1, y1, x2, y2))
            
            # Resize to fit
            aspect = cropped.height / cropped.width if cropped.width > 0 else 1
            new_w = width
            new_h = min(int(new_w * aspect), height)
            if new_h < 1:
                new_h = 1
            if new_w < 1:
                new_w = 1
                
            resized = cropped.resize((new_w, new_h), Image.Resampling.LANCZOS)
            return ImageTk.PhotoImage(resized)
        except Exception as e:
            print(f"[Thumbnail] Error: {e}")
            return None
            
    def _on_thumbnail_click(self, region, source: str):
        """Handle thumbnail click - scroll Source to region"""
        if self.on_row_select and region:
            # Create a minimal pair object for the callback
            class TempPair:
                def __init__(self, web_id, pdf_id, similarity):
                    self.web_id = web_id
                    self.pdf_id = pdf_id
                    self.similarity = similarity
                    
            if source == "web":
                pair = TempPair(region.area_code, None, region.similarity)
            else:
                pair = TempPair(None, region.area_code, region.similarity)
                
            self.on_row_select(pair.web_id, pair.pdf_id, pair)
            print(f"[Thumbnail] Clicked: {source} - {region.area_code}")
                
    def _on_row_click(self, row_widget, pair):
        """Handle row click - highlight and notify parent"""
        # Reset previous selection
        if self.selected_widget:
            try:
                idx = list(self.scroll_frame.winfo_children()).index(self.selected_widget)
                color = "#2D2D2D" if idx % 2 == 0 else "#252525"
                self.selected_widget.configure(fg_color=color)
            except:
                pass
            
        # Highlight new selection
        self.selected_widget = row_widget
        self.selected_indices = (pair.web_id, pair.pdf_id)
        row_widget.configure(fg_color="#444466")
        
        # Notify parent for Source panel sync
        if self.on_row_select:
            self.on_row_select(pair.web_id, pair.pdf_id, pair)
        
        print(f"[Spreadsheet] Selected: {pair.web_id} <-> {pair.pdf_id}")

    def get_selected_ids(self):
        """Return (web_id, pdf_id) or None"""
        return self.selected_indices

    def _on_export(self):
        """Export to Excel"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill
            from pathlib import Path
            from datetime import datetime
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Comparison"
            
            # Headers
            headers = ["No", "Web ID", "Web Text", "PDF ID", "PDF Text", "Sync %"]
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="4CAF50", fill_type="solid")
            
            # Data
            for i, pair in enumerate(self.sync_pairs, 2):
                web_region = self.web_map.get(pair.web_id)
                pdf_region = self.pdf_map.get(pair.pdf_id)
                
                ws.cell(row=i, column=1, value=i-1)
                ws.cell(row=i, column=2, value=pair.web_id)
                ws.cell(row=i, column=3, value=web_region.text[:500] if web_region else "")
                ws.cell(row=i, column=4, value=pair.pdf_id)
                ws.cell(row=i, column=5, value=pdf_region.text[:500] if pdf_region else "")
                ws.cell(row=i, column=6, value=f"{int(pair.similarity*100)}%")
            
            # Column widths
            ws.column_dimensions['C'].width = 60
            ws.column_dimensions['E'].width = 60
            
            # Save
            Path("./exports").mkdir(exist_ok=True)
            filename = f"comparison_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            output_path = f"./exports/{filename}"
            wb.save(output_path)
            
            print(f"Excel exported: {output_path}")
            import os
            os.startfile(output_path)
            
        except Exception as e:
            print(f"Export error: {e}")
            import tkinter.messagebox as mb
            mb.showerror("Export Error", str(e))
