"""
Ultimate Sync - Excel/Spreadsheet Exporter
パラグラフ比較結果をExcelに出力（プレビュー画像付き）
"""

import io
from pathlib import Path
from datetime import datetime
from typing import List, Optional, Tuple
from PIL import Image

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Fill, PatternFill, Alignment, Border, Side
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("⚠️ openpyxl not installed. Excel export will not be available.")

from app.core.paragraph_matcher import ParagraphEntry, SyncPair


class SyncExporter:
    """
    Sync結果をExcelファイルにエクスポート
    """
    
    def __init__(self, output_dir: str = "./exports"):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
    
    def _hex_to_rgb(self, hex_color: str) -> str:
        """#RRGGBB → RRGGBB"""
        return hex_color.lstrip('#')
    
    def _create_preview_image(
        self, 
        source_image: Image.Image, 
        rect: List[int],
        max_width: int = 150
    ) -> Optional[bytes]:
        """
        パラグラフ領域のプレビュー画像を生成
        """
        try:
            x1, y1, x2, y2 = rect
            # 境界チェック
            x1 = max(0, min(x1, source_image.width))
            x2 = max(0, min(x2, source_image.width))
            y1 = max(0, min(y1, source_image.height))
            y2 = max(0, min(y2, source_image.height))
            
            if x2 <= x1 or y2 <= y1:
                return None
            
            # 切り抜き
            cropped = source_image.crop((x1, y1, x2, y2))
            
            # リサイズ
            ratio = max_width / cropped.width if cropped.width > max_width else 1.0
            new_size = (int(cropped.width * ratio), int(cropped.height * ratio))
            if new_size[0] > 0 and new_size[1] > 0:
                cropped = cropped.resize(new_size, Image.Resampling.LANCZOS)
            
            # バイト列に変換
            img_bytes = io.BytesIO()
            cropped.save(img_bytes, format='PNG')
            img_bytes.seek(0)
            return img_bytes
        except Exception as e:
            print(f"[SyncExporter] Preview image error: {e}")
            return None
    
    def export_to_excel(
        self,
        web_paragraphs: List[ParagraphEntry],
        pdf_paragraphs: List[ParagraphEntry],
        sync_pairs: List[SyncPair],
        web_image: Optional[Image.Image] = None,
        pdf_image: Optional[Image.Image] = None,
        filename: str = None
    ) -> str:
        """
        Sync結果をExcelファイルに出力
        
        Returns:
            出力ファイルパス
        """
        if not OPENPYXL_AVAILABLE:
            raise RuntimeError("openpyxl is not installed. Run: pip install openpyxl")
        
        # ファイル名生成
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"sync_result_{timestamp}.xlsx"
        
        output_path = self.output_dir / filename
        
        # ワークブック作成
        wb = Workbook()
        
        # === Web シート ===
        ws_web = wb.active
        ws_web.title = "Web Paragraphs"
        self._create_paragraph_sheet(ws_web, web_paragraphs, web_image, "Web")
        
        # === PDF シート ===
        ws_pdf = wb.create_sheet("PDF Paragraphs")
        self._create_paragraph_sheet(ws_pdf, pdf_paragraphs, pdf_image, "PDF")
        
        # === Sync Summary シート ===
        ws_sync = wb.create_sheet("Sync Summary")
        self._create_sync_summary_sheet(ws_sync, sync_pairs, web_paragraphs, pdf_paragraphs)
        
        # 保存
        wb.save(output_path)
        print(f"✅ Excel出力完了: {output_path}")
        
        return str(output_path)
    
    def _create_paragraph_sheet(
        self, 
        ws, 
        paragraphs: List[ParagraphEntry],
        source_image: Optional[Image.Image],
        source_name: str
    ):
        """パラグラフ一覧シートを作成"""
        # ヘッダー
        headers = ["ID", "Page", "Text", "Sync ID", "Similarity", "Status"]
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # 列幅設定
        ws.column_dimensions['A'].width = 8   # ID
        ws.column_dimensions['B'].width = 6   # Page
        ws.column_dimensions['C'].width = 60  # Text
        ws.column_dimensions['D'].width = 10  # Sync ID
        ws.column_dimensions['E'].width = 12  # Similarity
        ws.column_dimensions['F'].width = 10  # Status
        
        # データ行
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row, p in enumerate(paragraphs, 2):
            # ID
            ws.cell(row=row, column=1, value=p.id)
            
            # Page
            ws.cell(row=row, column=2, value=p.page)
            
            # Text (最大200文字)
            text_preview = p.text[:200] + "..." if len(p.text) > 200 else p.text
            ws.cell(row=row, column=3, value=text_preview)
            
            # Sync ID
            ws.cell(row=row, column=4, value=p.sync_id or "-")
            
            # Similarity
            ws.cell(row=row, column=5, value=p.similarity_percent)
            
            # Status (色付き)
            status_cell = ws.cell(row=row, column=6)
            if p.similarity >= 0.5:
                status_cell.value = "🟢 Match"
                status_cell.fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
            elif p.similarity >= 0.3:
                status_cell.value = "🟡 Partial"
                status_cell.fill = PatternFill(start_color="FFE0B2", end_color="FFE0B2", fill_type="solid")
            elif p.sync_id:
                status_cell.value = "🔴 Low"
                status_cell.fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
            else:
                status_cell.value = "⚪ None"
            
            # 行全体にボーダー
            for col in range(1, 7):
                ws.cell(row=row, column=col).border = thin_border
    
    def _create_sync_summary_sheet(
        self, 
        ws, 
        sync_pairs: List[SyncPair],
        web_paragraphs: List[ParagraphEntry],
        pdf_paragraphs: List[ParagraphEntry]
    ):
        """Sync概要シートを作成"""
        # 統計
        total_web = len(web_paragraphs)
        total_pdf = len(pdf_paragraphs)
        high_matches = sum(1 for sp in sync_pairs if sp.similarity >= 0.5)
        mid_matches = sum(1 for sp in sync_pairs if 0.3 <= sp.similarity < 0.5)
        low_matches = sum(1 for sp in sync_pairs if sp.similarity < 0.3)
        unmatched_web = total_web - sum(1 for sp in sync_pairs)
        unmatched_pdf = total_pdf - sum(1 for sp in sync_pairs)
        
        # サマリー
        ws['A1'] = "Sync Analysis Summary"
        ws['A1'].font = Font(bold=True, size=14)
        
        summary_data = [
            ("", ""),
            ("Total Web Paragraphs", total_web),
            ("Total PDF Paragraphs", total_pdf),
            ("", ""),
            ("🟢 High Match (50%+)", high_matches),
            ("🟡 Partial Match (30-50%)", mid_matches),
            ("🔴 Low Match (<30%)", low_matches),
            ("⚪ Unmatched Web", unmatched_web),
            ("⚪ Unmatched PDF", unmatched_pdf),
        ]
        
        for row, (label, value) in enumerate(summary_data, 3):
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=2, value=value)
        
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15
        
        # マッチペア一覧
        start_row = len(summary_data) + 5
        ws.cell(row=start_row, column=1, value="Matched Pairs")
        ws.cell(row=start_row, column=1).font = Font(bold=True, size=12)
        
        headers = ["Web ID", "PDF ID", "Similarity", "Status"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row + 1, column=col, value=header)
            cell.font = Font(bold=True)
        
        for i, sp in enumerate(sync_pairs):
            row = start_row + 2 + i
            ws.cell(row=row, column=1, value=sp.web_id)
            ws.cell(row=row, column=2, value=sp.pdf_id)
            ws.cell(row=row, column=3, value=f"{sp.similarity * 100:.1f}%")
            
            status = "🟢" if sp.similarity >= 0.5 else "🟡" if sp.similarity >= 0.3 else "🔴"
            ws.cell(row=row, column=4, value=status)


def export_sync_results(
    web_paragraphs: List[ParagraphEntry],
    pdf_paragraphs: List[ParagraphEntry],
    sync_pairs: List[SyncPair],
    web_image: Optional[Image.Image] = None,
    pdf_image: Optional[Image.Image] = None,
    output_dir: str = "./exports"
) -> str:
    """
    便利関数: Sync結果をExcelにエクスポート
    """
    exporter = SyncExporter(output_dir)
    return exporter.export_to_excel(
        web_paragraphs, pdf_paragraphs, sync_pairs,
        web_image, pdf_image
    )
