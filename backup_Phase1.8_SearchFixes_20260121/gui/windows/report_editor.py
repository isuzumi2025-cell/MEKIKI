"""
レポートエディター
比較結果の編集、コメント加筆、Google Sheets/Excel出力
"""
import customtkinter as ctk
import tkinter as tk
from tkinter import ttk
from typing import Optional, Dict, List
from PIL import Image, ImageTk
import io
import json
from pathlib import Path


class ReportEditorWindow(ctk.CTkToplevel):
    """
    レポートエディター - 分離可能なウィンドウ
    比較結果の編集とエクスポート機能
    """
    
    def __init__(self, parent, **kwargs):
        super().__init__(parent, **kwargs)
        
        # データ
        self.report_items: List[Dict] = []
        
        # ウィンドウ設定
        self.title("📝 レポートエディター")
        self.geometry("1000x700")
        self.minsize(700, 500)
        
        self._build_ui()
    
    def _build_ui(self):
        """UI構築"""
        # ヘッダー
        header = ctk.CTkFrame(self, fg_color="#1A1A1A", height=50)
        header.pack(fill="x")
        header.pack_propagate(False)
        
        ctk.CTkLabel(
            header,
            text="📝 レポートエディター",
            font=("Meiryo", 16, "bold"),
            text_color="#4CAF50"
        ).pack(side="left", padx=15, pady=10)
        
        # エクスポートボタン
        export_frame = ctk.CTkFrame(header, fg_color="transparent")
        export_frame.pack(side="right", padx=10)
        
        ctk.CTkButton(
            export_frame, text="📊 Google Sheets", width=120, fg_color="#0F9D58",
            command=self._export_to_gsheets
        ).pack(side="left", padx=5)
        
        ctk.CTkButton(
            export_frame, text="📑 Excel", width=80, fg_color="#217346",
            command=self._export_to_excel
        ).pack(side="left", padx=5)
        
        ctk.CTkButton(
            export_frame, text="📄 JSON", width=70, fg_color="#616161",
            command=self._export_to_json
        ).pack(side="left", padx=5)
        
        # メインコンテンツ
        main = ctk.CTkFrame(self, fg_color="#2B2B2B")
        main.pack(fill="both", expand=True, padx=10, pady=10)
        
        # ツールバー
        toolbar = ctk.CTkFrame(main, fg_color="#383838", height=40, corner_radius=10)
        toolbar.pack(fill="x", padx=5, pady=5)
        toolbar.pack_propagate(False)
        
        ctk.CTkButton(
            toolbar, text="➕ 項目追加", width=100, fg_color="#FF6F00",
            command=self._add_item
        ).pack(side="left", padx=10, pady=5)
        
        ctk.CTkButton(
            toolbar, text="🗑️ 選択削除", width=100, fg_color="#F44336",
            command=self._delete_selected
        ).pack(side="left", padx=5)
        
        ctk.CTkLabel(
            toolbar, text="📋 レポート項目一覧", font=("Meiryo", 11)
        ).pack(side="right", padx=15)
        
        # 項目リスト
        list_frame = ctk.CTkFrame(main, fg_color="#2D2D2D", corner_radius=10)
        list_frame.pack(fill="both", expand=True, padx=5, pady=5)
        
        # Treeview スタイル
        style = ttk.Style()
        style.configure("Report.Treeview",
                        background="#2D2D2D",
                        foreground="white",
                        fieldbackground="#2D2D2D",
                        rowheight=60)
        style.configure("Report.Treeview.Heading",
                        background="#383838",
                        foreground="white")
        
        # Treeview
        columns = ("thumbnail", "source", "text", "sync", "comment")
        self.tree = ttk.Treeview(
            list_frame,
            style="Report.Treeview",
            columns=columns,
            show="headings",
            selectmode="extended"
        )
        
        self.tree.heading("thumbnail", text="プレビュー")
        self.tree.heading("source", text="ソース")
        self.tree.heading("text", text="抽出テキスト")
        self.tree.heading("sync", text="Sync率")
        self.tree.heading("comment", text="コメント")
        
        self.tree.column("thumbnail", width=80, anchor="center")
        self.tree.column("source", width=100, anchor="center")
        self.tree.column("text", width=300, anchor="w")
        self.tree.column("sync", width=80, anchor="center")
        self.tree.column("comment", width=250, anchor="w")
        
        # スクロールバー
        scrollbar = ttk.Scrollbar(list_frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=scrollbar.set)
        
        self.tree.pack(side="left", fill="both", expand=True, padx=(10, 0), pady=10)
        scrollbar.pack(side="right", fill="y", pady=10, padx=(0, 5))
        
        # ダブルクリックで編集
        self.tree.bind("<Double-1>", self._on_item_double_click)
        
        # 下部: 編集パネル
        self.edit_panel = ctk.CTkFrame(main, fg_color="#383838", corner_radius=10, height=150)
        self.edit_panel.pack(fill="x", padx=5, pady=5)
        self.edit_panel.pack_propagate(False)
        
        edit_header = ctk.CTkFrame(self.edit_panel, fg_color="transparent")
        edit_header.pack(fill="x", padx=10, pady=5)
        
        ctk.CTkLabel(
            edit_header, text="✏️ コメント編集", font=("Meiryo", 11, "bold")
        ).pack(side="left")
        
        ctk.CTkButton(
            edit_header, text="💾 保存", width=70, fg_color="#4CAF50",
            command=self._save_comment
        ).pack(side="right")
        
        self.comment_entry = ctk.CTkTextbox(
            self.edit_panel, font=("Meiryo", 11), fg_color="#2D2D2D", height=80
        )
        self.comment_entry.pack(fill="x", padx=10, pady=5)
        
        # ステータスバー
        self.status_bar = ctk.CTkFrame(self, height=25, fg_color="#1A1A1A")
        self.status_bar.pack(fill="x", side="bottom")
        self.status_bar.pack_propagate(False)
        
        self.status_label = ctk.CTkLabel(
            self.status_bar,
            text="レポート項目: 0件",
            font=("Meiryo", 10),
            text_color="gray"
        )
        self.status_label.pack(side="left", padx=10)
    
    def add_comparison_result(self, web_text: str, pdf_text: str, sync_rate: float,
                               web_image: Optional[Image.Image] = None,
                               pdf_image: Optional[Image.Image] = None):
        """比較結果を追加"""
        item = {
            "id": len(self.report_items) + 1,
            "source": "Web vs PDF",
            "web_text": web_text[:50] + "..." if len(web_text) > 50 else web_text,
            "pdf_text": pdf_text[:50] + "..." if len(pdf_text) > 50 else pdf_text,
            "sync_rate": sync_rate,
            "comment": "",
            "web_image": web_image,
            "pdf_image": pdf_image
        }
        
        self.report_items.append(item)
        self._refresh_list()
    
    def _refresh_list(self):
        """リストを更新"""
        for item in self.tree.get_children():
            self.tree.delete(item)
        
        for item in self.report_items:
            sync = f"{item['sync_rate']:.0f}%"
            self.tree.insert("", "end", values=(
                "📷",  # サムネイル（実際は画像表示が複雑）
                item["source"],
                item["web_text"],
                sync,
                item["comment"]
            ), tags=(f"item_{item['id']}",))
        
        self.status_label.configure(text=f"レポート項目: {len(self.report_items)}件")
    
    def _add_item(self):
        """項目を追加"""
        # ダイアログ表示 (簡易版)
        dialog = ctk.CTkInputDialog(
            text="テキストを入力:",
            title="項目追加"
        )
        text = dialog.get_input()
        
        if text:
            self.add_comparison_result(text, text, 100.0)
    
    def _delete_selected(self):
        """選択項目を削除"""
        selected = self.tree.selection()
        if not selected:
            return
        
        for item_id in selected:
            self.tree.delete(item_id)
        
        # report_itemsも同期 (簡易版)
        self.report_items = [
            item for i, item in enumerate(self.report_items)
            if f"item_{item['id']}" not in [self.tree.item(s)['tags'][0] for s in selected if self.tree.exists(s)]
        ]
        
        self._refresh_list()
    
    def _on_item_double_click(self, event):
        """項目ダブルクリック"""
        selection = self.tree.selection()
        if selection:
            item = self.tree.item(selection[0])
            # コメント欄に既存コメントをロード
            self.comment_entry.delete("1.0", "end")
            self.comment_entry.insert("1.0", item['values'][4])
    
    def _save_comment(self):
        """コメントを保存"""
        selection = self.tree.selection()
        if not selection:
            return
        
        comment = self.comment_entry.get("1.0", "end").strip()
        
        # Treeview更新
        for sel in selection:
            values = list(self.tree.item(sel)['values'])
            values[4] = comment
            self.tree.item(sel, values=values)
        
        self.status_label.configure(text="💾 コメント保存完了")
    
    def _export_to_gsheets(self):
        """Google Sheetsにエクスポート"""
        try:
            from app.core.gsheet_exporter import export_to_gsheet
            
            data = self._prepare_export_data()
            result = export_to_gsheet(data)
            
            from tkinter import messagebox
            messagebox.showinfo("完了", f"Google Sheetsにエクスポートしました\n{result}")
            
        except ImportError:
            from tkinter import messagebox
            messagebox.showwarning("未設定", "Google Sheets APIが設定されていません\nservice_account.jsonを確認してください")
        except Exception as e:
            from tkinter import messagebox
            messagebox.showerror("エラー", f"エクスポートエラー: {e}")
    
    def _export_to_excel(self):
        """Excelにエクスポート"""
        from tkinter import filedialog, messagebox
        
        file_path = filedialog.asksaveasfilename(
            title="Excelファイルを保存",
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")]
        )
        
        if not file_path:
            return
        
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "比較レポート"
            
            # ヘッダー
            headers = ["No", "ソース", "抽出テキスト", "Sync率", "コメント"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
            
            # データ
            for row, item in enumerate(self.report_items, 2):
                ws.cell(row=row, column=1, value=item["id"])
                ws.cell(row=row, column=2, value=item["source"])
                ws.cell(row=row, column=3, value=item["web_text"])
                ws.cell(row=row, column=4, value=f"{item['sync_rate']:.0f}%")
                ws.cell(row=row, column=5, value=item["comment"])
            
            wb.save(file_path)
            messagebox.showinfo("完了", f"Excelに保存しました\n{file_path}")
            
        except ImportError:
            messagebox.showwarning("未設定", "openpyxlがインストールされていません\npip install openpyxl")
        except Exception as e:
            messagebox.showerror("エラー", f"保存エラー: {e}")
    
    def _export_to_json(self):
        """JSONにエクスポート"""
        from tkinter import filedialog, messagebox
        
        file_path = filedialog.asksaveasfilename(
            title="JSONファイルを保存",
            defaultextension=".json",
            filetypes=[("JSON", "*.json")]
        )
        
        if not file_path:
            return
        
        try:
            data = self._prepare_export_data()
            
            with open(file_path, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
            
            messagebox.showinfo("完了", f"JSONに保存しました\n{file_path}")
            
        except Exception as e:
            messagebox.showerror("エラー", f"保存エラー: {e}")
    
    def _prepare_export_data(self) -> List[Dict]:
        """エクスポート用データを準備"""
        return [
            {
                "id": item["id"],
                "source": item["source"],
                "web_text": item["web_text"],
                "pdf_text": item.get("pdf_text", ""),
                "sync_rate": item["sync_rate"],
                "comment": item["comment"]
            }
            for item in self.report_items
        ]


class ReportEditorFrame(ctk.CTkFrame):
    """レポートエディター - 埋め込み用フレーム版（機能強化）"""
    
    def __init__(self, parent, **kwargs):
        super().__init__(parent, fg_color="transparent", **kwargs)
        
        self.parent_app = parent.winfo_toplevel()
        self.report_items: List[Dict] = []
        
        self._build_ui()
        
        # 定期更新
        self.after(2000, self._auto_refresh)
    
    def _build_ui(self):
        """UI構築"""
        # ヘッダー
        header = ctk.CTkFrame(self, fg_color="#1A1A1A", height=50)
        header.pack(fill="x")
        header.pack_propagate(False)
        
        ctk.CTkLabel(
            header,
            text="📝 レポートエディター",
            font=("Meiryo", 16, "bold"),
            text_color="#4CAF50"
        ).pack(side="left", padx=15, pady=10)
        
        # サマリー表示
        self.summary_label = ctk.CTkLabel(
            header, text="0件", font=("Meiryo", 11), text_color="gray"
        )
        self.summary_label.pack(side="left", padx=10)
        
        ctk.CTkButton(
            header, text="↗️ 別ウィンドウで開く",
            command=self._open_window, fg_color="#616161"
        ).pack(side="right", padx=15)
        
        ctk.CTkButton(
            header, text="📑 Excel出力", width=80,
            command=self._quick_export, fg_color="#217346"
        ).pack(side="right", padx=5)
        
        # メインコンテンツ
        main = ctk.CTkFrame(self, fg_color="#2D2D2D")
        main.pack(fill="both", expand=True, padx=10, pady=10)
        
        # クイックサマリーパネル
        summary_panel = ctk.CTkFrame(main, fg_color="#383838", corner_radius=10)
        summary_panel.pack(fill="x", padx=10, pady=10)
        
        stats_frame = ctk.CTkFrame(summary_panel, fg_color="transparent")
        stats_frame.pack(fill="x", padx=15, pady=10)
        
        # 統計カード
        self._create_stat_card(stats_frame, "📊 合計項目", "items_count", 0)
        self._create_stat_card(stats_frame, "✅ 高Sync率", "high_sync", 1)
        self._create_stat_card(stats_frame, "⚠️ 要確認", "low_sync", 2)
        self._create_stat_card(stats_frame, "📝 コメント済", "commented", 3)
        
        # レポートリスト (スクロール可能)
        list_frame = ctk.CTkFrame(main, fg_color="#2D2D2D")
        list_frame.pack(fill="both", expand=True, padx=10, pady=(0, 10))
        
        list_header = ctk.CTkFrame(list_frame, fg_color="#424242", height=30, corner_radius=8)
        list_header.pack(fill="x", pady=(0, 5))
        list_header.pack_propagate(False)
        
        ctk.CTkLabel(
            list_header, text="最新レポート項目", font=("Meiryo", 10, "bold")
        ).pack(side="left", padx=10, pady=5)
        
        self.list_scroll = ctk.CTkScrollableFrame(list_frame, fg_color="transparent")
        self.list_scroll.pack(fill="both", expand=True)
        
        # 初期メッセージ
        self.empty_label = ctk.CTkLabel(
            self.list_scroll,
            text="校正結果をレポートに追加すると\nここに一覧表示されます",
            font=("Meiryo", 11), text_color="gray"
        )
        self.empty_label.pack(expand=True, pady=30)
    
    def _create_stat_card(self, parent, title: str, key: str, col: int):
        """統計カード作成"""
        parent.grid_columnconfigure(col, weight=1)
        
        card = ctk.CTkFrame(parent, fg_color="#2D2D2D", corner_radius=8)
        card.grid(row=0, column=col, padx=5, sticky="ew")
        
        ctk.CTkLabel(
            card, text=title, font=("Meiryo", 9), text_color="gray"
        ).pack(pady=(8, 2))
        
        value_label = ctk.CTkLabel(
            card, text="--", font=("Meiryo", 18, "bold"), text_color="white"
        )
        value_label.pack(pady=(0, 8))
        
        setattr(self, f"stat_{key}", value_label)
    
    def _auto_refresh(self):
        """自動更新"""
        self._update_stats()
        self.after(5000, self._auto_refresh)
    
    def _update_stats(self):
        """統計更新"""
        count = len(self.report_items)
        high = sum(1 for item in self.report_items if item.get('sync_rate', 0) >= 95)
        low = sum(1 for item in self.report_items if item.get('sync_rate', 0) < 70)
        commented = sum(1 for item in self.report_items if item.get('comment', ''))
        
        self.stat_items_count.configure(text=str(count))
        self.stat_high_sync.configure(text=str(high))
        self.stat_low_sync.configure(text=str(low))
        self.stat_commented.configure(text=str(commented))
        self.summary_label.configure(text=f"{count}件")
    
    def add_item(self, web_text: str, pdf_text: str, sync_rate: float, comment: str = ""):
        """レポート項目追加"""
        self.report_items.append({
            'id': len(self.report_items) + 1,
            'web_text': web_text[:100],
            'pdf_text': pdf_text[:100],
            'sync_rate': sync_rate,
            'comment': comment
        })
        
        self.empty_label.pack_forget()
        self._add_item_widget(self.report_items[-1])
        self._update_stats()
    
    def _add_item_widget(self, item: Dict):
        """項目ウィジェット追加"""
        frame = ctk.CTkFrame(self.list_scroll, fg_color="#383838", corner_radius=8)
        frame.pack(fill="x", pady=3)
        
        # 左: Sync率
        sync = item.get('sync_rate', 0)
        color = "#4CAF50" if sync >= 95 else "#FF9800" if sync >= 70 else "#F44336"
        
        ctk.CTkLabel(
            frame, text=f"{sync:.0f}%", font=("Meiryo", 12, "bold"),
            text_color=color, width=50
        ).pack(side="left", padx=10, pady=8)
        
        # 中央: テキストプレビュー
        text_preview = item.get('web_text', '')[:50] + "..."
        ctk.CTkLabel(
            frame, text=text_preview, font=("Meiryo", 10),
            text_color="white", anchor="w"
        ).pack(side="left", fill="x", expand=True, padx=5)
        
        # 右: コメントアイコン
        if item.get('comment'):
            ctk.CTkLabel(
                frame, text="💬", font=("Meiryo", 12)
            ).pack(side="right", padx=10)
    
    def _quick_export(self):
        """クイックExcel出力"""
        if not self.report_items:
            from tkinter import messagebox
            messagebox.showinfo("情報", "レポート項目がありません")
            return
        
        window = ReportEditorWindow(self.winfo_toplevel())
        for item in self.report_items:
            window.add_comparison_result(
                item.get('web_text', ''),
                item.get('pdf_text', ''),
                item.get('sync_rate', 0)
            )
        window._export_to_excel()
    
    def _open_window(self):
        """別ウィンドウで開く"""
        window = ReportEditorWindow(self.winfo_toplevel())
        for item in self.report_items:
            window.add_comparison_result(
                item.get('web_text', ''),
                item.get('pdf_text', ''),
                item.get('sync_rate', 0)
            )
        window.focus()
