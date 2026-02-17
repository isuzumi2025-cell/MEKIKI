"""
OCR Metadata Exporter
検出されたテキストとエリア座標をCSV/Excelに出力

Phase 2: パイプライン再設計
"""

import csv
import os
from datetime import datetime
from typing import List, Dict, Any, Optional
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False


class MetadataExporter:
    """OCRメタデータをCSV/Excelに出力"""
    
    COLUMNS = [
        'ID', 'Source', 'Page', 'X1', 'Y1', 'X2', 'Y2', 
        'Width', 'Height', 'TextLength', 'Text'
    ]
    
    def __init__(self, output_dir: str = "./exports"):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
    
    def _generate_filename(self, prefix: str, extension: str) -> str:
        """タイムスタンプ付きファイル名を生成"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        return f"{prefix}_{timestamp}.{extension}"
    
    def _cluster_to_row(self, cluster: Dict, source: str, index: int) -> Dict:
        """クラスタを行データに変換"""
        prefix = "W" if source == "web" else "P"
        rect = cluster.get('rect', [0, 0, 0, 0])
        text = cluster.get('text', '')
        
        return {
            'ID': f'{prefix}-{index+1:03d}',
            'Source': source,
            'Page': cluster.get('page', 1),
            'X1': rect[0] if len(rect) > 0 else 0,
            'Y1': rect[1] if len(rect) > 1 else 0,
            'X2': rect[2] if len(rect) > 2 else 0,
            'Y2': rect[3] if len(rect) > 3 else 0,
            'Width': (rect[2] - rect[0]) if len(rect) > 2 else 0,
            'Height': (rect[3] - rect[1]) if len(rect) > 3 else 0,
            'TextLength': len(text),
            'Text': text[:500]  # 最大500文字
        }
    
    def export_to_csv(
        self, 
        web_clusters: List[Dict], 
        pdf_clusters: List[Dict],
        filename: Optional[str] = None
    ) -> str:
        """CSVに出力"""
        if filename is None:
            filename = self._generate_filename("metadata", "csv")
        
        output_path = self.output_dir / filename
        
        rows = []
        for i, c in enumerate(web_clusters):
            rows.append(self._cluster_to_row(c, 'web', i))
        for i, c in enumerate(pdf_clusters):
            rows.append(self._cluster_to_row(c, 'pdf', i))
        
        with open(output_path, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.DictWriter(f, fieldnames=self.COLUMNS)
            writer.writeheader()
            writer.writerows(rows)
        
        print(f"[MetadataExporter] CSV exported: {output_path}")
        return str(output_path)
    
    def export_to_excel(
        self, 
        web_clusters: List[Dict], 
        pdf_clusters: List[Dict],
        filename: Optional[str] = None
    ) -> Optional[str]:
        """Excelに出力"""
        if not EXCEL_AVAILABLE:
            print("[MetadataExporter] openpyxl not available, skipping Excel export")
            return None
        
        if filename is None:
            filename = self._generate_filename("metadata", "xlsx")
        
        output_path = self.output_dir / filename
        
        wb = openpyxl.Workbook()
        
        # Webシート
        ws_web = wb.active
        ws_web.title = "Web"
        self._write_sheet(ws_web, web_clusters, 'web')
        
        # PDFシート
        ws_pdf = wb.create_sheet("PDF")
        self._write_sheet(ws_pdf, pdf_clusters, 'pdf')
        
        # 統合シート
        ws_all = wb.create_sheet("All")
        all_clusters = []
        for i, c in enumerate(web_clusters):
            all_clusters.append(self._cluster_to_row(c, 'web', i))
        for i, c in enumerate(pdf_clusters):
            all_clusters.append(self._cluster_to_row(c, 'pdf', i))
        self._write_rows_to_sheet(ws_all, all_clusters)
        
        wb.save(output_path)
        print(f"[MetadataExporter] Excel exported: {output_path}")
        return str(output_path)
    
    def _write_sheet(self, ws, clusters: List[Dict], source: str):
        """シートにデータを書き込み"""
        rows = [self._cluster_to_row(c, source, i) for i, c in enumerate(clusters)]
        self._write_rows_to_sheet(ws, rows)
    
    def _write_rows_to_sheet(self, ws, rows: List[Dict]):
        """行データをシートに書き込み"""
        # ヘッダー
        header_fill = PatternFill(start_color="4CAF50", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col, header in enumerate(self.COLUMNS, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
        
        # データ
        for row_idx, row_data in enumerate(rows, 2):
            for col, header in enumerate(self.COLUMNS, 1):
                ws.cell(row=row_idx, column=col, value=row_data.get(header, ''))
        
        # カラム幅調整
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 6
        ws.column_dimensions['K'].width = 60


def export_ocr_metadata(web_clusters: List[Dict], pdf_clusters: List[Dict], output_dir: str = "./exports") -> Dict[str, str]:
    """
    OCRメタデータをCSV/Excelに出力するユーティリティ関数
    
    Returns:
        {'csv': 'path/to/csv', 'excel': 'path/to/excel'}
    """
    exporter = MetadataExporter(output_dir)
    
    result = {}
    result['csv'] = exporter.export_to_csv(web_clusters, pdf_clusters)
    
    excel_path = exporter.export_to_excel(web_clusters, pdf_clusters)
    if excel_path:
        result['excel'] = excel_path
    
    return result


def export_comparison_results(match_pairs: List[Dict], output_dir: str = "./exports") -> Optional[str]:
    """
    Phase 4: 比較結果をExcelに出力
    
    RUNBOOKフォーマット:
    | # | Web ID | Web Text | ⇔ | PDF Text | PDF ID | Score |
    
    Args:
        match_pairs: [{'web_id': 'W-001', 'pdf_id': 'P-003', 'common': '...', 'common_len': 33, 'web_text': '...', 'pdf_text': '...'}]
        output_dir: 出力ディレクトリ
    
    Returns:
        出力ファイルパス
    """
    if not EXCEL_AVAILABLE:
        print("[ComparisonExporter] openpyxl not available")
        return None
    
    from datetime import datetime
    from pathlib import Path
    
    output_path = Path(output_dir)
    output_path.mkdir(parents=True, exist_ok=True)
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = output_path / f"comparison_{timestamp}.xlsx"
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Comparison"
    
    # ヘッダー (RUNBOOKフォーマット)
    headers = ['#', 'Web ID', 'Web Text', '⇔', 'PDF Text', 'PDF ID', 'CommonLen', 'Common Substring']
    header_fill = PatternFill(start_color="4CAF50", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # データ
    for i, m in enumerate(match_pairs, 1):
        ws.cell(row=i+1, column=1, value=i)
        ws.cell(row=i+1, column=2, value=m.get('web_id', ''))
        ws.cell(row=i+1, column=3, value=m.get('web_text', '')[:200])
        ws.cell(row=i+1, column=4, value='✓')
        ws.cell(row=i+1, column=5, value=m.get('pdf_text', '')[:200])
        ws.cell(row=i+1, column=6, value=m.get('pdf_id', ''))
        ws.cell(row=i+1, column=7, value=m.get('common_len', 0))
        ws.cell(row=i+1, column=8, value=m.get('common', '')[:100])
    
    # カラム幅調整
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 8
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['D'].width = 3
    ws.column_dimensions['E'].width = 50
    ws.column_dimensions['F'].width = 8
    ws.column_dimensions['G'].width = 10
    ws.column_dimensions['H'].width = 40
    
    wb.save(filename)
    print(f"[ComparisonExporter] Excel exported: {filename}")
    return str(filename)

