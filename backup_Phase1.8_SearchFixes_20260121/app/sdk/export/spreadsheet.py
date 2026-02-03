"""
Spreadsheet Exporter
検版結果をExcel/CSVにエクスポート

Created: 2026-01-11
"""
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Any, Optional
from dataclasses import dataclass
import csv

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


# 列定義
COLUMN_DEFINITIONS = [
    ("run_id", "実行ID", 15),
    ("page_left", "左ページ", 10),
    ("page_right", "右ページ", 10),
    ("element_kind", "種別", 12),
    ("left_text_norm", "左テキスト", 40),
    ("right_text_norm", "右テキスト", 40),
    ("field_types", "フィールド", 15),
    ("diff_type", "差分種別", 15),
    ("severity", "重大度", 12),
    ("risk_reason", "リスク理由", 30),
    ("score_total", "スコア", 10),
    ("evidence_left_crop", "左証拠", 25),
    ("evidence_right_crop", "右証拠", 25),
    ("evidence_overlay", "オーバーレイ", 25),
    ("status", "ステータス", 12),
    ("reviewer", "担当者", 15),
    ("comment", "コメント", 40),
    ("updated_at", "更新日時", 20),
]

# 重大度の色
SEVERITY_COLORS = {
    "CRITICAL": "FF0000",  # 赤
    "MAJOR": "FFA500",     # オレンジ
    "MINOR": "FFFF00",     # 黄
    "INFO": "FFFFFF",      # 白
}


@dataclass
class ExportResult:
    """エクスポート結果"""
    file_path: str
    row_count: int
    created_at: datetime
    error: Optional[str] = None


class SpreadsheetExporter:
    """
    スプレッドシートエクスポーター
    Excel/CSV形式で検版結果を出力
    """
    
    def __init__(
        self,
        storage_path: Optional[str] = None,
        template_path: Optional[str] = None
    ):
        """
        初期化
        
        Args:
            storage_path: 保存先ディレクトリ
            template_path: テンプレートファイルパス
        """
        self.storage_path = Path(storage_path) if storage_path else Path("storage/runs")
        self.template_path = Path(template_path) if template_path else None
    
    def export_xlsx(
        self,
        issues: List[Dict[str, Any]],
        run_id: str,
        filename: Optional[str] = None
    ) -> ExportResult:
        """
        Excel形式でエクスポート
        
        Args:
            issues: Issue辞書のリスト
            run_id: 実行ID
            filename: ファイル名（省略時は自動生成）
            
        Returns:
            ExportResult
        """
        if not OPENPYXL_AVAILABLE:
            return ExportResult(
                file_path="",
                row_count=0,
                created_at=datetime.now(),
                error="openpyxl is not installed"
            )
        
        try:
            # 保存先
            save_dir = self.storage_path / run_id / "exports"
            save_dir.mkdir(parents=True, exist_ok=True)
            
            filename = filename or f"proofing_result_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            file_path = save_dir / filename
            
            # ワークブック作成
            wb = Workbook()
            ws = wb.active
            ws.title = "検版結果"
            
            # ヘッダー設定
            self._setup_header(ws)
            
            # データ行追加
            for i, issue in enumerate(issues, start=2):
                self._add_row(ws, i, issue)
            
            # 列幅調整
            self._adjust_column_widths(ws)
            
            # 保存
            wb.save(str(file_path))
            
            return ExportResult(
                file_path=str(file_path),
                row_count=len(issues),
                created_at=datetime.now()
            )
            
        except Exception as e:
            return ExportResult(
                file_path="",
                row_count=0,
                created_at=datetime.now(),
                error=str(e)
            )
    
    def export_csv(
        self,
        issues: List[Dict[str, Any]],
        run_id: str,
        filename: Optional[str] = None
    ) -> ExportResult:
        """
        CSV形式でエクスポート
        """
        try:
            save_dir = self.storage_path / run_id / "exports"
            save_dir.mkdir(parents=True, exist_ok=True)
            
            filename = filename or f"proofing_result_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
            file_path = save_dir / filename
            
            with open(file_path, 'w', newline='', encoding='utf-8-sig') as f:
                writer = csv.DictWriter(
                    f,
                    fieldnames=[col[0] for col in COLUMN_DEFINITIONS]
                )
                
                # ヘッダー（日本語ラベル）
                writer.writerow({col[0]: col[1] for col in COLUMN_DEFINITIONS})
                
                # データ
                for issue in issues:
                    row = self._prepare_row_data(issue)
                    writer.writerow(row)
            
            return ExportResult(
                file_path=str(file_path),
                row_count=len(issues),
                created_at=datetime.now()
            )
            
        except Exception as e:
            return ExportResult(
                file_path="",
                row_count=0,
                created_at=datetime.now(),
                error=str(e)
            )
    
    def update_existing(
        self,
        file_path: str,
        new_issues: List[Dict[str, Any]]
    ) -> ExportResult:
        """
        既存ファイルを更新（追加）
        """
        if not OPENPYXL_AVAILABLE:
            return ExportResult(
                file_path="",
                row_count=0,
                created_at=datetime.now(),
                error="openpyxl is not installed"
            )
        
        try:
            wb = load_workbook(file_path)
            ws = wb.active
            
            # 現在の最終行を取得
            start_row = ws.max_row + 1
            
            # 新規行追加
            for i, issue in enumerate(new_issues):
                self._add_row(ws, start_row + i, issue)
            
            wb.save(file_path)
            
            return ExportResult(
                file_path=file_path,
                row_count=len(new_issues),
                created_at=datetime.now()
            )
            
        except Exception as e:
            return ExportResult(
                file_path="",
                row_count=0,
                created_at=datetime.now(),
                error=str(e)
            )
    
    def _setup_header(self, ws):
        """ヘッダー行を設定"""
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        
        for col_idx, (key, label, width) in enumerate(COLUMN_DEFINITIONS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=label)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
    
    def _add_row(self, ws, row_idx: int, issue: Dict[str, Any]):
        """データ行を追加"""
        row_data = self._prepare_row_data(issue)
        
        for col_idx, (key, _, _) in enumerate(COLUMN_DEFINITIONS, start=1):
            value = row_data.get(key, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            
            # 重大度セルに色付け
            if key == "severity":
                color = SEVERITY_COLORS.get(value, "FFFFFF")
                cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    
    def _prepare_row_data(self, issue: Dict[str, Any]) -> Dict[str, str]:
        """Issue辞書からrow用データを準備"""
        return {
            "run_id": issue.get("run_id", ""),
            "page_left": str(issue.get("page_left", "")),
            "page_right": str(issue.get("page_right", "")),
            "element_kind": issue.get("element_kind", ""),
            "left_text_norm": self._truncate(issue.get("left_text_norm", ""), 100),
            "right_text_norm": self._truncate(issue.get("right_text_norm", ""), 100),
            "field_types": ",".join(issue.get("field_types", [])) if isinstance(issue.get("field_types"), list) else str(issue.get("field_types", "")),
            "diff_type": issue.get("diff_type", ""),
            "severity": issue.get("severity", "INFO"),
            "risk_reason": ",".join(issue.get("risk_reason", [])) if isinstance(issue.get("risk_reason"), list) else str(issue.get("risk_reason", "")),
            "score_total": f"{issue.get('score_total', 0):.2f}" if isinstance(issue.get("score_total"), (int, float)) else "",
            "evidence_left_crop": issue.get("evidence_left_crop", ""),
            "evidence_right_crop": issue.get("evidence_right_crop", ""),
            "evidence_overlay": issue.get("evidence_overlay", ""),
            "status": issue.get("status", "OPEN"),
            "reviewer": issue.get("reviewer", ""),
            "comment": issue.get("comment", ""),
            "updated_at": issue.get("updated_at", datetime.now().isoformat()),
        }
    
    def _adjust_column_widths(self, ws):
        """列幅を調整"""
        for col_idx, (_, _, width) in enumerate(COLUMN_DEFINITIONS, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width
    
    def _truncate(self, text: str, max_len: int) -> str:
        """テキストを切り詰め"""
        if len(text) > max_len:
            return text[:max_len - 3] + "..."
        return text


# ========== Convenience Function ==========

def export_issues(
    issues: List[Dict[str, Any]],
    run_id: str,
    format: str = "xlsx"
) -> ExportResult:
    """
    Issue一覧をエクスポート（簡易インターフェース）
    
    Args:
        issues: Issue辞書のリスト
        run_id: 実行ID
        format: 出力形式 ("xlsx" or "csv")
    """
    exporter = SpreadsheetExporter()
    
    if format == "csv":
        return exporter.export_csv(issues, run_id)
    else:
        return exporter.export_xlsx(issues, run_id)
