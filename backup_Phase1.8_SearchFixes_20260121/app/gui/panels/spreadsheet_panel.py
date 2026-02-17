import customtkinter as ctk
import tkinter as tk
from typing import List, Any, Callable, Optional
from PIL import Image, ImageTk
import difflib
from datetime import datetime
from pathlib import Path

# ãƒ•ã‚¡ã‚¤ãƒ«ãƒ­ã‚®ãƒ³ã‚°è¨­å®š
LOG_FILE = Path(__file__).parent.parent.parent.parent / "lcs_diagnostic.log"

def log_diagnostic(msg: str):
    """è¨ºæ–­ãƒ­ã‚°ã‚’ãƒ•ã‚¡ã‚¤ãƒ«ã«å‡ºåŠ›"""
    try:
        timestamp = datetime.now().strftime("%H:%M:%S")
        with open(LOG_FILE, "a", encoding="utf-8") as f:
            f.write(f"[{timestamp}] {msg}\n")
        print(msg)  # ã‚³ãƒ³ã‚½ãƒ¼ãƒ«ã«ã‚‚å‡ºåŠ›
    except Exception as e:
        print(f"[LOG ERROR] {e}: {msg}")

# ãƒ¢ã‚¸ãƒ¥ãƒ¼ãƒ«èª­ã¿è¾¼ã¿ç¢ºèª
log_diagnostic(f"=== SpreadsheetPanel MODULE LOADED ===")
log_diagnostic(f"LOG_FILE path: {LOG_FILE}")



class SpreadsheetPanel(ctk.CTkFrame):
    """
    Live Comparison Sheet (Virtual Spreadsheet)
    - Thumbnails below ID (click to jump to Source)
    - Score display with color coding
    - Full text display
    """

    # å®šæ•°
    LCS_FONT_SIZE = 13
    MAX_TEXT_LENGTH = 10000
    THUMB_WIDTH = 60
    THUMB_HEIGHT = 80
    ROW_HEIGHT = 140
    SCORE_WIDTH = 70
    ID_COLUMN_WIDTH = 100  # Web ID/Thumb ã¨ PDF ID/Thumb ã®å¹…ã‚’çµ±ä¸€

    def __init__(self, parent, on_row_select: Optional[Callable] = None, **kwargs):
        super().__init__(parent, **kwargs)
        log_diagnostic("[SpreadsheetPanel] __init__ called")

        self.sync_pairs = []
        self.web_map = {}
        self.pdf_map = {}
        self.web_image = None
        self.pdf_image = None
        self.selected_indices = None
        self.selected_widget = None
        self.on_row_select = on_row_select
        self._thumbnail_refs = []

        # Virtual list state
        self._visible_rows = {}  # {index: row_widget}
        self._visible_range = (0, 0)  # (start_index, end_index)
        self._rows_per_page = 15  # Number of visible rows to render
        self._scroll_update_pending = False

        self._build_ui()
        log_diagnostic("[SpreadsheetPanel] UI built successfully")

    def set_on_row_select(self, callback: Callable):
        """Set callback for row selection"""
        self.on_row_select = callback

    def set_images(self, web_image: Image.Image, pdf_image: Image.Image):
        """Set source images for thumbnail generation"""
        self.web_image = web_image
        self.pdf_image = pdf_image

    def _build_ui(self):
        # 1. Toolbar
        toolbar = ctk.CTkFrame(self, height=30, fg_color="#333333")
        toolbar.pack(fill="x", side="top")

        ctk.CTkLabel(toolbar, text="Live Comparison Sheet", font=("Meiryo", 12, "bold")).pack(side="left", padx=10)

        self.stats_label = ctk.CTkLabel(toolbar, text="Web: - | PDF: - | Match: -", font=("Meiryo", 11))
        self.stats_label.pack(side="left", padx=20)

        self.export_btn = ctk.CTkButton(
            toolbar,
            text="Excel Export",
            width=100,
            height=24,
            state="disabled",
            command=self._on_export
        )
        self.export_btn.pack(side="right", padx=5)

        # 2. Table Header
        header_frame = ctk.CTkFrame(self, height=28, fg_color="#2B2B2B")
        header_frame.pack(fill="x", side="top", pady=(1, 0))

        # Header columns - LEGACY ORDER: Score â†’ Web ID â†’ Web Text â†’ Arrow â†’ PDF Text â†’ PDF ID
        headers = [
            ("Score", 50),
            ("Web ID / Thumb", 100),
            ("Web Text", 0),
            ("", 30),
            ("PDF Text", 0),
            ("PDF ID / Thumb", 100),
        ]

        for text, width in headers:
            if width > 0:
                lbl = ctk.CTkLabel(header_frame, text=text, width=width, font=("Meiryo", 9, "bold"))
            else:
                lbl = ctk.CTkLabel(header_frame, text=text, font=("Meiryo", 9, "bold"))
            lbl.pack(side="left", padx=1, pady=4)
            if width == 0:
                lbl.pack_configure(expand=True, fill="x")

        # 3. Scrollable content area
        self.scroll_frame = ctk.CTkScrollableFrame(self, fg_color="#1E1E1E", corner_radius=0)
        self.scroll_frame.pack(fill="both", expand=True)

        # Bind scroll events for virtual list
        self.scroll_frame.bind("<Configure>", self._on_scroll_configure)
        self.scroll_frame._parent_canvas.bind("<MouseWheel>", self._on_scroll_event)
        self.scroll_frame._parent_canvas.bind("<Button-4>", self._on_scroll_event)  # Linux scroll up
        self.scroll_frame._parent_canvas.bind("<Button-5>", self._on_scroll_event)  # Linux scroll down

    def update_data(self, sync_pairs: List[Any], web_regions: List[Any], pdf_regions: List[Any],
                    web_image=None, pdf_image=None):
        """Update data and refresh display"""
        self.sync_pairs = sync_pairs

        # === DIAGNOSTIC LOG START ===
        log_diagnostic("="*60)
        log_diagnostic("[LCS update_data] DIAGNOSTIC")
        log_diagnostic(f"  sync_pairs count: {len(sync_pairs)}")
        log_diagnostic(f"  web_regions count: {len(web_regions)}")
        log_diagnostic(f"  pdf_regions count: {len(pdf_regions)}")
        log_diagnostic(f"  web_image: {web_image.size if web_image else 'None'}")
        log_diagnostic(f"  pdf_image: {pdf_image.size if pdf_image else 'None'}")

        if sync_pairs:
            sp = sync_pairs[0]
            log_diagnostic("  First SyncPair:")
            log_diagnostic(f"    web_id: {sp.web_id}")
            log_diagnostic(f"    pdf_id: {sp.pdf_id}")
            log_diagnostic(f"    web_bbox: {getattr(sp, 'web_bbox', 'N/A')}")
            log_diagnostic(f"    pdf_bbox: {getattr(sp, 'pdf_bbox', 'N/A')}")
            log_diagnostic(f"    web_text: {repr(getattr(sp, 'web_text', 'N/A')[:50]) if getattr(sp, 'web_text', None) else 'None'}")
            log_diagnostic(f"    pdf_text: {repr(getattr(sp, 'pdf_text', 'N/A')[:50]) if getattr(sp, 'pdf_text', None) else 'None'}")

        if web_regions:
            r = web_regions[0]
            log_diagnostic("  First web_region:")
            log_diagnostic(f"    area_code: {getattr(r, 'area_code', 'N/A')}")
            log_diagnostic(f"    id: {getattr(r, 'id', 'N/A')}")
            log_diagnostic(f"    rect: {getattr(r, 'rect', 'N/A')}")

        if pdf_regions:
            r = pdf_regions[0]
            log_diagnostic("  First pdf_region:")
            log_diagnostic(f"    area_code: {getattr(r, 'area_code', 'N/A')}")
            log_diagnostic(f"    id: {getattr(r, 'id', 'N/A')}")
            log_diagnostic(f"    rect: {getattr(r, 'rect', 'N/A')}")
        log_diagnostic("="*60)
        # === DIAGNOSTIC LOG END ===

        # EditableRegionã¯area_codeã‚’ã‚­ãƒ¼ã«ä½¿ç”¨ï¼ˆSyncPair.web_id/pdf_idã¨ä¸€è‡´ã•ã›ã‚‹ï¼‰
        self.web_map = {}
        self.pdf_map = {}

        for r in web_regions:
            key = getattr(r, 'area_code', None) or getattr(r, 'id', None)
            if key is not None:
                self.web_map[str(key)] = r

        for r in pdf_regions:
            key = getattr(r, 'area_code', None) or getattr(r, 'id', None)
            if key is not None:
                self.pdf_map[str(key)] = r

        # === KEY VERIFICATION ===
        if sync_pairs and self.web_map:
            sp = sync_pairs[0]
            found_web = self.web_map.get(sp.web_id)
            found_pdf = self.pdf_map.get(sp.pdf_id)
            log_diagnostic(f"[KEY CHECK] web_id={sp.web_id} -> found={found_web is not None}")
            log_diagnostic(f"[KEY CHECK] pdf_id={sp.pdf_id} -> found={found_pdf is not None}")
            if not found_web:
                log_diagnostic(f"[KEY CHECK] web_map keys (first 5): {list(self.web_map.keys())[:5]}")
            if not found_pdf:
                log_diagnostic(f"[KEY CHECK] pdf_map keys (first 5): {list(self.pdf_map.keys())[:5]}")

        if web_image:
            self.web_image = web_image
        if pdf_image:
            self.pdf_image = pdf_image

        # Stats update
        total_web = len(web_regions)
        total_pdf = len(pdf_regions)
        matched = sum(1 for p in sync_pairs if p.similarity >= 0.25)
        self.stats_label.configure(text=f"Web: {total_web} | PDF: {total_pdf} | Match: {matched}")

        self._refresh_rows()

        if sync_pairs:
            self.export_btn.configure(state="normal")

    def _refresh_rows(self):
        """Clear and rebuild rows using virtual list (only visible rows)"""
        log_diagnostic(f"[_refresh_rows] Starting VIRTUAL: {len(self.sync_pairs)} pairs total")
        log_diagnostic(f"[_refresh_rows] web_image: {self.web_image.size if self.web_image else 'None'}")
        log_diagnostic(f"[_refresh_rows] pdf_image: {self.pdf_image.size if self.pdf_image else 'None'}")

        # Clear all existing widgets
        for widget in self.scroll_frame.winfo_children():
            widget.destroy()
        self._visible_rows = {}
        self._thumbnail_refs = []

        # Calculate initial visible range (first page)
        total_rows = len(self.sync_pairs)
        if total_rows == 0:
            return

        # Set scrollregion based on total rows
        total_height = total_rows * (self.ROW_HEIGHT + 1)  # +1 for pady
        self.scroll_frame._parent_canvas.configure(scrollregion=(0, 0, 800, total_height))

        # Render initial visible rows
        end_index = min(self._rows_per_page, total_rows)
        self._visible_range = (0, end_index)
        self._render_visible_rows()

        log_diagnostic(f"[_refresh_rows] Done VIRTUAL: Rendered {end_index} of {total_rows} rows")

        # Force UI update to ensure widgets are displayed
        self.scroll_frame.update_idletasks()

    def _render_visible_rows(self):
        """Render only the rows in the visible range"""
        start_idx, end_idx = self._visible_range

        # Clear existing visible rows
        for widget in self.scroll_frame.winfo_children():
            widget.destroy()
        self._visible_rows = {}
        self._thumbnail_refs = []

        # Add top spacer for scrolled-past rows
        if start_idx > 0:
            top_spacer_height = start_idx * (self.ROW_HEIGHT + 1)
            top_spacer = ctk.CTkFrame(self.scroll_frame, height=top_spacer_height, fg_color="#1E1E1E")
            top_spacer.pack(fill="x")
            top_spacer.pack_propagate(False)

        # Render visible rows
        for i in range(start_idx, end_idx):
            if i < len(self.sync_pairs):
                pair = self.sync_pairs[i]
                row_widget = self._create_row(i, pair)
                self._visible_rows[i] = row_widget

        # Add bottom spacer for remaining rows
        remaining_rows = len(self.sync_pairs) - end_idx
        if remaining_rows > 0:
            bottom_spacer_height = remaining_rows * (self.ROW_HEIGHT + 1)
            bottom_spacer = ctk.CTkFrame(self.scroll_frame, height=bottom_spacer_height, fg_color="#1E1E1E")
            bottom_spacer.pack(fill="x")
            bottom_spacer.pack_propagate(False)

        log_diagnostic(f"[Virtual] Rendered rows {start_idx} to {end_idx-1} (total: {len(self.sync_pairs)})")

    def _on_scroll_event(self, event):
        """Handle scroll events to trigger virtual list updates"""
        if not self._scroll_update_pending:
            self._scroll_update_pending = True
            self.after(50, self._update_visible_range)  # Debounce 50ms

    def _on_scroll_configure(self, event):
        """Handle configure events"""
        if not self._scroll_update_pending:
            self._scroll_update_pending = True
            self.after(50, self._update_visible_range)

    def _update_visible_range(self):
        """Calculate and update visible row range based on scroll position"""
        self._scroll_update_pending = False

        if not self.sync_pairs:
            return

        try:
            # Get current scroll position (0.0 to 1.0)
            canvas = self.scroll_frame._parent_canvas
            yview = canvas.yview()
            scroll_top = yview[0]  # Top of visible area (0.0 = top, 1.0 = bottom)

            # Calculate visible row indices
            total_rows = len(self.sync_pairs)
            total_height = total_rows * (self.ROW_HEIGHT + 1)

            # Current scroll position in pixels
            scroll_y_px = scroll_top * total_height

            # Calculate visible row range with buffer (render extra rows above/below)
            buffer_rows = 5
            start_idx = max(0, int(scroll_y_px / (self.ROW_HEIGHT + 1)) - buffer_rows)
            end_idx = min(total_rows, start_idx + self._rows_per_page + (buffer_rows * 2))

            # Only update if range changed significantly
            old_start, old_end = self._visible_range
            if abs(start_idx - old_start) > 3 or abs(end_idx - old_end) > 3:
                self._visible_range = (start_idx, end_idx)
                self._render_visible_rows()

        except Exception as e:
            log_diagnostic(f"[Virtual] Scroll update error: {e}")

    def _create_row(self, index: int, pair):
        """Create a single row with thumbnails below ID"""
        row_bg = "#2D2D2D" if index % 2 == 0 else "#252525"

        row = ctk.CTkFrame(self.scroll_frame, fg_color=row_bg, corner_radius=0, height=self.ROW_HEIGHT)
        row.pack(fill="x", pady=1)
        row.pack_propagate(False)

        # Get regions from map
        web_region = self.web_map.get(pair.web_id)
        pdf_region = self.pdf_map.get(pair.pdf_id)

        # === ROW DIAGNOSTIC (first 3 rows only) ===
        if index < 3:
            log_diagnostic(f"[Row {index}] web_id={pair.web_id}, pdf_id={pair.pdf_id}")
            log_diagnostic(f"  web_region found: {web_region is not None}")
            log_diagnostic(f"  pdf_region found: {pdf_region is not None}")
            log_diagnostic(f"  pair.web_bbox: {getattr(pair, 'web_bbox', None)}")
            log_diagnostic(f"  pair.pdf_bbox: {getattr(pair, 'pdf_bbox', None)}")
            log_diagnostic(f"  self.web_image: {self.web_image.size if self.web_image else 'None'}")
            log_diagnostic(f"  self.pdf_image: {self.pdf_image.size if self.pdf_image else 'None'}")

        # Get text (prefer pair data, fallback to region)
        w_txt = getattr(pair, 'web_text', '') or (web_region.text if web_region else "")
        p_txt = getattr(pair, 'pdf_text', '') or (pdf_region.text if pdf_region else "")

        # Get bbox (prefer pair data, fallback to region)
        web_bbox = getattr(pair, 'web_bbox', None)
        if not web_bbox and web_region and hasattr(web_region, 'rect'):
            web_bbox = web_region.rect

        pdf_bbox = getattr(pair, 'pdf_bbox', None)
        if not pdf_bbox and pdf_region and hasattr(pdf_region, 'rect'):
            pdf_bbox = pdf_region.rect

        # === BBOX DIAGNOSTIC (first 3 rows only) ===
        if index < 3:
            log_diagnostic(f"  final web_bbox: {web_bbox}")
            log_diagnostic(f"  final pdf_bbox: {pdf_bbox}")

        # Score calculation
        sim_percent = int(pair.similarity * 100)
        if sim_percent >= 80:
            score_color, score_bg = "#4CAF50", "#1B3D1B"
        elif sim_percent >= 50:
            score_color, score_bg = "#FFC107", "#3D3D1B"
        elif sim_percent >= 30:
            score_color, score_bg = "#FF9800", "#3D2D1B"
        else:
            score_color, score_bg = "#F44336", "#3D1B1B"

        # === LEGACY PACK ORDER: All LEFT, in sequence ===
        # Score â†’ Web ID â†’ Web Text â†’ Arrow â†’ PDF Text â†’ PDF ID

        # 1. Score (LEFT)
        score_frame = ctk.CTkFrame(row, fg_color=score_bg, width=50)
        score_frame.pack(side="left", fill="y", padx=1)
        score_frame.pack_propagate(False)
        score_label = ctk.CTkLabel(score_frame, text=f"{sim_percent}%", text_color=score_color,
                     font=("Arial", 10, "bold"))
        score_label.pack(expand=True, fill="both")

        # 2. Web ID + Thumbnail (LEFT)
        web_id_frame = ctk.CTkFrame(row, fg_color=row_bg, width=100)
        web_id_frame.pack(side="left", fill="y", padx=1)
        web_id_frame.pack_propagate(False)

        ctk.CTkLabel(web_id_frame, text=pair.web_id or "-", text_color="#4CAF50", font=("Meiryo", 8)).pack(pady=(2, 0))

        web_thumb = self._create_thumbnail(self.web_image, web_bbox)
        if web_thumb:
            self._thumbnail_refs.append(web_thumb)
            web_thumb_label = tk.Label(web_id_frame, image=web_thumb, bg=row_bg, cursor="hand2")
            web_thumb_label.pack(pady=2)
            web_thumb_label.bind("<Button-1>", lambda e, b=web_bbox: self._on_thumbnail_click(b, "web", pair))

        # 3. Web Text (LEFT, expand)
        web_text_frame = ctk.CTkFrame(row, fg_color=row_bg)
        web_text_frame.pack(side="left", fill="both", expand=True, padx=2)

        web_text_widget = tk.Text(web_text_frame, bg=row_bg, fg="#E0E0E0", relief="flat",
                          font=("Meiryo", 9), wrap="word", height=5, width=25)
        web_text_widget.pack(fill="both", expand=True, padx=2, pady=2)

        # 4. Arrow (LEFT)
        ctk.CTkLabel(row, text="<>", width=30, text_color="#666666").pack(side="left", padx=1)

        # 5. PDF Text (LEFT, expand)
        pdf_text_frame = ctk.CTkFrame(row, fg_color=row_bg)
        pdf_text_frame.pack(side="left", fill="both", expand=True, padx=2)

        pdf_text_widget = tk.Text(pdf_text_frame, bg=row_bg, fg="#E0E0E0", relief="flat",
                          font=("Meiryo", 9), wrap="word", height=5, width=25)
        pdf_text_widget.pack(fill="both", expand=True, padx=2, pady=2)
        
        # â˜… Diff Highlight é©ç”¨ (ä¸€è‡´=ã‚°ãƒ¬ãƒ¼ã€å·®åˆ†=èµ¤/é’)
        self._apply_diff_highlight(web_text_widget, pdf_text_widget, w_txt[:200], p_txt[:200])

        # 6. PDF ID + Thumbnail (LEFT - last)
        pdf_id_frame = ctk.CTkFrame(row, fg_color=row_bg, width=100)
        pdf_id_frame.pack(side="left", fill="y", padx=1)
        pdf_id_frame.pack_propagate(False)

        ctk.CTkLabel(pdf_id_frame, text=pair.pdf_id or "-", text_color="#2196F3", font=("Meiryo", 8)).pack(pady=(2, 0))

        pdf_thumb = self._create_thumbnail(self.pdf_image, pdf_bbox)
        if pdf_thumb:
            self._thumbnail_refs.append(pdf_thumb)
            pdf_thumb_label = tk.Label(pdf_id_frame, image=pdf_thumb, bg=row_bg, cursor="hand2")
            pdf_thumb_label.pack(pady=2)
            pdf_thumb_label.bind("<Button-1>", lambda e, b=pdf_bbox: self._on_thumbnail_click(b, "pdf", pair))

        # 7. Action Buttons (Similar/Match)
        action_frame = ctk.CTkFrame(row, fg_color=row_bg, width=70)
        action_frame.pack(side="left", fill="y", padx=2)
        action_frame.pack_propagate(False)
        
        # Similar Search Button
        similar_btn = ctk.CTkButton(
            action_frame,
            text="ğŸ”",
            width=28,
            height=24,
            fg_color="#424242",
            hover_color="#616161",
            font=("Segoe UI Emoji", 12),
            command=lambda p=pair: self._on_similar_search(p)
        )
        similar_btn.pack(pady=(20, 2))
        
        # Match Search Button
        match_btn = ctk.CTkButton(
            action_frame,
            text="ğŸ¯",
            width=28,
            height=24,
            fg_color="#424242",
            hover_color="#616161",
            font=("Segoe UI Emoji", 12),
            command=lambda p=pair: self._on_match_search(p)
        )
        match_btn.pack(pady=2)

        # Row click binding
        row.bind("<Button-1>", lambda e, p=pair, w=row: self._on_row_click(w, p))
        for widget in [score_frame, web_id_frame, pdf_id_frame]:
            widget.bind("<Button-1>", lambda e, p=pair, w=row: self._on_row_click(w, p))

        return row

    def _create_thumbnail(self, source_image, bbox):
        """Create a thumbnail from the source image and bbox [x1, y1, x2, y2]"""
        if not source_image:
            log_diagnostic("[Thumb] FAIL: source_image is None")
            return None
        if not bbox:
            log_diagnostic("[Thumb] FAIL: bbox is None")
            return None

        try:
            x1, y1, x2, y2 = bbox
            x1 = max(0, int(x1))
            y1 = max(0, int(y1))
            x2 = min(source_image.width, int(x2))
            y2 = min(source_image.height, int(y2))

            if x2 <= x1 or y2 <= y1:
                log_diagnostic(f"[Thumb] FAIL: Invalid bbox after clamp ({x1},{y1},{x2},{y2}) img={source_image.size}")
                return None

            cropped = source_image.crop((x1, y1, x2, y2))
            log_diagnostic(f"[Thumb] OK: cropped {cropped.size} from bbox ({x1},{y1},{x2},{y2})")

            # Resize to fit within THUMB_WIDTH x THUMB_HEIGHT
            aspect = cropped.height / cropped.width if cropped.width > 0 else 1
            if aspect > self.THUMB_HEIGHT / self.THUMB_WIDTH:
                new_h = self.THUMB_HEIGHT
                new_w = max(1, int(new_h / aspect))
            else:
                new_w = self.THUMB_WIDTH
                new_h = max(1, int(new_w * aspect))

            resized = cropped.resize((new_w, new_h), Image.Resampling.LANCZOS)
            # ãƒ¬ã‚¬ã‚·ãƒ¼æ–¹å¼: ImageTk.PhotoImage ã‚’ä½¿ç”¨
            photo_img = ImageTk.PhotoImage(resized)
            log_diagnostic(f"[Thumb] SUCCESS: PhotoImage created ({new_w}x{new_h})")
            return photo_img

        except Exception as e:
            log_diagnostic(f"[Thumb] EXCEPTION: {e}")
            import traceback
            log_diagnostic(traceback.format_exc())
            return None

    def _on_thumbnail_click(self, bbox, source: str, pair):
        """Handle thumbnail click - notify parent to highlight region"""
        if self.on_row_select and bbox:
            self.on_row_select(pair.web_id, pair.pdf_id, pair)

    def _on_row_click(self, row_widget, pair):
        """Handle row click - highlight and notify parent"""
        # Reset previous selection
        if self.selected_widget:
            try:
                idx = list(self.scroll_frame.winfo_children()).index(self.selected_widget)
                color = "#2D2D2D" if idx % 2 == 0 else "#252525"
                self.selected_widget.configure(fg_color=color)
            except:
                pass

        # Highlight new selection
        self.selected_widget = row_widget
        self.selected_indices = (pair.web_id, pair.pdf_id)
        row_widget.configure(fg_color="#444466")

        # Notify parent
        if self.on_row_select:
            self.on_row_select(pair.web_id, pair.pdf_id, pair)

    def get_selected_ids(self):
        """Return (web_id, pdf_id) or None"""
        return self.selected_indices
    
    def set_on_similar_search(self, callback: Callable):
        """Set callback for Similar Search button"""
        self.on_similar_search = callback
    
    def set_on_match_search(self, callback: Callable):
        """Set callback for Match Search button"""
        self.on_match_search = callback
    
    def _on_similar_search(self, pair):
        """Handle Similar Search button click"""
        log_diagnostic(f"[SimilarSearch] Triggered for pair: web={pair.web_id}, pdf={pair.pdf_id}")
        if hasattr(self, 'on_similar_search') and self.on_similar_search:
            self.on_similar_search(pair)
        else:
            log_diagnostic("[SimilarSearch] No callback set")
            print("ğŸ” é¡ä¼¼æ¤œç´¢: ã‚³ãƒ¼ãƒ«ãƒãƒƒã‚¯æœªè¨­å®š")
    
    def _on_match_search(self, pair):
        """Handle Match Search button click"""
        log_diagnostic(f"[MatchSearch] Triggered for pair: web={pair.web_id}, pdf={pair.pdf_id}")
        if hasattr(self, 'on_match_search') and self.on_match_search:
            self.on_match_search(pair)
        else:
            log_diagnostic("[MatchSearch] No callback set")
            print("ğŸ¯ ãƒãƒƒãƒæ¤œç´¢: ã‚³ãƒ¼ãƒ«ãƒãƒƒã‚¯æœªè¨­å®š")
    
    def _apply_diff_highlight(self, w_widget, p_widget, t1: str, t2: str):
        """ãƒ†ã‚­ã‚¹ãƒˆå·®åˆ†ã‚’è‰²åˆ†ã‘ã—ã¦è¡¨ç¤º (Gemini SemanticDiff + LCSãƒ•ã‚©ãƒ¼ãƒ«ãƒãƒƒã‚¯)"""
        import re
        
        # Tags: ä¸€è‡´=ç™½ã€å·®åˆ†ã®ã¿=ç·‘
        w_widget.tag_config("normal", foreground="#E0E0E0")
        w_widget.tag_config("diff", foreground="#4CAF50", background="#1A3D1A")
        p_widget.tag_config("normal", foreground="#E0E0E0")
        p_widget.tag_config("diff", foreground="#4CAF50", background="#1A3D1A")
        
        if not t1 and not t2:
            w_widget.configure(state="disabled")
            p_widget.configure(state="disabled")
            return
        
        # ãƒ†ã‚­ã‚¹ãƒˆæ­£è¦åŒ– (æ‹¬å¼§ãƒ»æ¥å°¾è¾é™¤å»)
        def normalize_for_matching(text):
            """ãƒãƒƒãƒãƒ³ã‚°ç”¨ã«æ­£è¦åŒ– - æ‹¬å¼§ãƒ»æ¥å°¾è¾ã‚’é™¤å»"""
            if not text:
                return ""
            # æ‹¬å¼§é¡ã‚’é™¤å»
            text = re.sub(r'[ã€ã€‘ã€Œã€ã€ã€ï¼ˆï¼‰()\[\]ã€Šã€‹ã€ˆã€‰ãƒ»ï¼š:ã€ã€‚]', '', text)
            # ç©ºç™½ãƒ»æ”¹è¡Œã‚’é™¤å»
            text = re.sub(r'\s+', '', text)
            return text[:200]
        
        # ãƒãƒƒãƒãƒ³ã‚°ç”¨æ­£è¦åŒ–ãƒ†ã‚­ã‚¹ãƒˆ
        match_t1 = normalize_for_matching(t1)
        match_t2 = normalize_for_matching(t2)
        
        # è¡¨ç¤ºç”¨ãƒ†ã‚­ã‚¹ãƒˆ (ç©ºç™½ã®ã¿é™¤å»)
        def clean(text):
            if not text:
                return ""
            text = re.sub(r'\s+', '', text)
            return text[:200]
        
        clean_t1 = clean(t1)
        clean_t2 = clean(t2)
        
        # â˜… Gemini SemanticDiff ã‚’è©¦è¡Œ
        matches = []
        try:
            from app.sdk.similarity.semantic_diff import get_semantic_diff
            semantic = get_semantic_diff()
            result = semantic.analyze(t1[:300], t2[:300])
            matches = result.matches if result else []
            if matches:
                log_diagnostic(f"[SemanticDiff] Found {len(matches)} matches via Gemini")
        except Exception as e:
            log_diagnostic(f"[SemanticDiff] Fallback to LCS: {e}")
        
        # GeminiãƒãƒƒãƒãŒã‚ã‚Œã°ä½¿ç”¨ã€ãªã‘ã‚Œã°LCSãƒ•ã‚©ãƒ¼ãƒ«ãƒãƒƒã‚¯
        if matches:
            # Geminiãƒãƒƒãƒã‚’ä½¿ç”¨
            def highlight_with_matches(widget, text, match_list):
                text = text[:200]
                clean_text = re.sub(r'\s+', '', text)
                
                # ãƒãƒƒãƒä½ç½®ã‚’åé›†
                match_ranges = []
                for phrase in sorted(match_list, key=len, reverse=True):
                    if len(phrase) < 3:
                        continue
                    pattern = re.escape(phrase)
                    for m in re.finditer(pattern, clean_text, re.IGNORECASE):
                        match_ranges.append((m.start(), m.end()))
                
                # ç¯„å›²ã‚’ã‚½ãƒ¼ãƒˆãƒ»ãƒãƒ¼ã‚¸
                match_ranges.sort()
                merged = []
                for start, end in match_ranges:
                    if merged and start <= merged[-1][1]:
                        merged[-1] = (merged[-1][0], max(merged[-1][1], end))
                    else:
                        merged.append((start, end))
                
                # è‰²åˆ†ã‘è¡¨ç¤º
                pos = 0
                for start, end in merged:
                    if pos < start:
                        widget.insert("end", clean_text[pos:start], "diff")
                    widget.insert("end", clean_text[start:end], "normal")
                    pos = end
                if pos < len(clean_text):
                    widget.insert("end", clean_text[pos:], "diff")
            
            highlight_with_matches(w_widget, t1, matches)
            highlight_with_matches(p_widget, t2, matches)
        else:
            # â˜… N-gram + LCS ãƒã‚¤ãƒ–ãƒªãƒƒãƒ‰ãƒãƒƒãƒãƒ³ã‚° (æ­£è¦åŒ–ãƒ†ã‚­ã‚¹ãƒˆä½¿ç”¨)
            # Step 1: N-gramã§å…±é€šãƒ•ãƒ¬ãƒ¼ã‚ºã‚’æ¤œå‡º (4æ–‡å­—ã‚¹ãƒ©ã‚¤ãƒ‡ã‚£ãƒ³ã‚°ã‚¦ã‚£ãƒ³ãƒ‰ã‚¦)
            NGRAM_SIZE = 4  # 5â†’4 ã«ç¸®å°ã—ã¦ç²’åº¦ã‚’ç´°ã‹ã
            
            def extract_ngrams(text, n=NGRAM_SIZE):
                """ã‚¹ãƒ©ã‚¤ãƒ‡ã‚£ãƒ³ã‚°ã‚¦ã‚£ãƒ³ãƒ‰ã‚¦ã§N-gramæŠ½å‡º"""
                ngrams = {}
                for i in range(len(text) - n + 1):
                    gram = text[i:i+n]
                    if gram not in ngrams:
                        ngrams[gram] = []
                    ngrams[gram].append(i)
                return ngrams
            
            # â˜… æ­£è¦åŒ–ãƒ†ã‚­ã‚¹ãƒˆã§N-gramæŠ½å‡º (æ‹¬å¼§ãªã—)
            t1_ngrams = extract_ngrams(match_t1, NGRAM_SIZE)
            t2_ngrams = extract_ngrams(match_t2, NGRAM_SIZE)
            
            # å…±é€šN-gramã‚’æ¤œå‡º
            common_ngrams = set(t1_ngrams.keys()) & set(t2_ngrams.keys())
            log_diagnostic(f"[N-gram] Found {len(common_ngrams)} common {NGRAM_SIZE}-grams")
            
            # Step 2: æ­£è¦åŒ–ãƒ†ã‚­ã‚¹ãƒˆã®ãƒãƒƒãƒç¯„å›²ã‚’æ§‹ç¯‰
            def build_match_ranges(ngrams, common_set, n=NGRAM_SIZE):
                """å…±é€šN-gramã®ä½ç½®ã‹ã‚‰ãƒãƒƒãƒç¯„å›²ã‚’æ§‹ç¯‰"""
                ranges = []
                for gram in common_set:
                    if gram in ngrams:
                        for pos in ngrams[gram]:
                            ranges.append((pos, pos + n))
                
                # ç¯„å›²ã‚’ã‚½ãƒ¼ãƒˆãƒ»ãƒãƒ¼ã‚¸
                ranges.sort()
                merged = []
                for start, end in ranges:
                    if merged and start <= merged[-1][1]:
                        merged[-1] = (merged[-1][0], max(merged[-1][1], end))
                    else:
                        merged.append((start, end))
                return merged
            
            # â˜… Step 3: æ˜ç¤ºçš„ãªå…±é€šéƒ¨åˆ†æ–‡å­—åˆ—æ¤œç´¢ (N-gramæ¼ã‚Œã‚’è£œå®Œ)
            # æ­£è¦åŒ–ãƒ†ã‚­ã‚¹ãƒˆã§è¦‹ã¤ã‹ã£ãŸå…±é€šéƒ¨åˆ†ã‚’ã€è¡¨ç¤ºãƒ†ã‚­ã‚¹ãƒˆã§ä½ç½®ç‰¹å®š
            
            def find_common_substrings(text1, text2, min_len=4):
                """ä¸¡æ–¹ã®ãƒ†ã‚­ã‚¹ãƒˆã«å­˜åœ¨ã™ã‚‹éƒ¨åˆ†æ–‡å­—åˆ—ã‚’æ¤œå‡º"""
                common = set()
                # text1ã®å…¨éƒ¨åˆ†æ–‡å­—åˆ—ã‚’ãƒã‚§ãƒƒã‚¯
                for length in range(min_len, min(30, len(text1)) + 1):
                    for i in range(len(text1) - length + 1):
                        substr = text1[i:i+length]
                        if substr in text2:
                            common.add(substr)
                return common
            
            # æ­£è¦åŒ–ãƒ†ã‚­ã‚¹ãƒˆã§å…±é€šéƒ¨åˆ†ã‚’æ¤œå‡º
            common_substrings = find_common_substrings(match_t1, match_t2, 4)
            log_diagnostic(f"[Substring] Found {len(common_substrings)} common substrings (4+ chars)")
            
            # é•·ã„ã‚‚ã®ã‚’å„ªå…ˆã—ã¦ãƒãƒƒãƒä½ç½®ã‚’ç‰¹å®š
            def find_positions_in_text(text, substrings):
                """ãƒ†ã‚­ã‚¹ãƒˆå†…ã®ãƒãƒƒãƒä½ç½®ã‚’æ¤œå‡º"""
                ranges = []
                # é•·ã„é †ã«ã‚½ãƒ¼ãƒˆï¼ˆé‡è¤‡ã‚’é¿ã‘ã‚‹ãŸã‚ï¼‰
                for substr in sorted(substrings, key=len, reverse=True):
                    if len(substr) >= 4:
                        pos = 0
                        while True:
                            idx = text.find(substr, pos)
                            if idx == -1:
                                break
                            ranges.append((idx, idx + len(substr)))
                            pos = idx + 1
                return ranges
            
            # è¡¨ç¤ºãƒ†ã‚­ã‚¹ãƒˆã§ãƒãƒƒãƒä½ç½®ã‚’æ¤œå‡º
            t1_ranges = find_positions_in_text(clean_t1, common_substrings)
            t2_ranges = find_positions_in_text(clean_t2, common_substrings)
            
            # Step 4: LCSã‚‚è¿½åŠ  (çŸ­ã„ãƒãƒƒãƒã‚’è£œå®Œ)
            matcher = difflib.SequenceMatcher(None, clean_t1, clean_t2)
            for m in matcher.get_matching_blocks():
                if m.size >= 3:
                    t1_ranges.append((m.a, m.a + m.size))
                    t2_ranges.append((m.b, m.b + m.size))
            
            # ç¯„å›²ã‚’å†ãƒãƒ¼ã‚¸
            def merge_ranges(ranges):
                ranges.sort()
                merged = []
                for start, end in ranges:
                    if merged and start <= merged[-1][1]:
                        merged[-1] = (merged[-1][0], max(merged[-1][1], end))
                    else:
                        merged.append((start, end))
                return merged
            
            t1_merged = merge_ranges(t1_ranges)
            t2_merged = merge_ranges(t2_ranges)
            
            # t1è¡¨ç¤º
            pos = 0
            for start, end in t1_merged:
                if pos < start:
                    w_widget.insert("end", clean_t1[pos:start], "diff")
                w_widget.insert("end", clean_t1[start:end], "normal")
                pos = end
            if pos < len(clean_t1):
                w_widget.insert("end", clean_t1[pos:], "diff")
            
            # t2è¡¨ç¤º
            pos = 0
            for start, end in t2_merged:
                if pos < start:
                    p_widget.insert("end", clean_t2[pos:start], "diff")
                p_widget.insert("end", clean_t2[start:end], "normal")
                pos = end
            if pos < len(clean_t2):
                p_widget.insert("end", clean_t2[pos:], "diff")
        
        w_widget.configure(state="disabled")
        p_widget.configure(state="disabled")

    def _on_export(self):
        """Export to Excel"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill
            from pathlib import Path
            from datetime import datetime

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Comparison"

            headers = ["No", "Web ID", "Web Text", "PDF ID", "PDF Text", "Sync %"]
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="4CAF50", fill_type="solid")

            for i, pair in enumerate(self.sync_pairs, 2):
                web_region = self.web_map.get(pair.web_id)
                pdf_region = self.pdf_map.get(pair.pdf_id)

                w_text = getattr(pair, 'web_text', '') or (web_region.text[:500] if web_region else "")
                p_text = getattr(pair, 'pdf_text', '') or (pdf_region.text[:500] if pdf_region else "")

                ws.cell(row=i, column=1, value=i-1)
                ws.cell(row=i, column=2, value=pair.web_id)
                ws.cell(row=i, column=3, value=w_text[:500])
                ws.cell(row=i, column=4, value=pair.pdf_id)
                ws.cell(row=i, column=5, value=p_text[:500])
                ws.cell(row=i, column=6, value=f"{int(pair.similarity*100)}%")

            ws.column_dimensions['C'].width = 60
            ws.column_dimensions['E'].width = 60

            Path("./exports").mkdir(exist_ok=True)
            filename = f"comparison_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            output_path = f"./exports/{filename}"
            wb.save(output_path)

            print(f"Excel exported: {output_path}")
            import os
            os.startfile(output_path)

        except Exception as e:
            print(f"Export error: {e}")
            import tkinter.messagebox as mb
            mb.showerror("Export Error", str(e))
